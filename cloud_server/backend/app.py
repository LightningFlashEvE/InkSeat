#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESP32 E-Paper Cloud Server - Flask Backend
Deep-sleep + HTTP Pull Architecture (无MQTT版本)

设备通过HTTP拉取更新，服务器持久化保存图片数据
"""

import os
import json
import time
import threading
import hashlib
import secrets
import re
import base64
import mimetypes
import shutil
import math
import zipfile
from datetime import datetime, timedelta, timezone
from functools import wraps
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit
from zoneinfo import ZoneInfo

import numpy as np
import requests
from PIL import Image

from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from pymongo import MongoClient, ReturnDocument
from pymongo.errors import DuplicateKeyError
from werkzeug.security import check_password_hash, generate_password_hash
import tempfile
import io

from config import Config
from db_indexes import ensure_all_indexes
from six_color_epd import process_e6_image_from_base64
from template_renderer import render_template_with_preview, _fetch_weather, _fetch_quote, _encode_epd_string

# ==================== Flask 应用初始化 ====================
app = Flask(__name__)
app.config.from_object(Config)
if Config.CORS_ORIGINS:
    CORS(
        app,
        resources={r'/api/*': {'origins': Config.CORS_ORIGINS}},
        allow_headers=[
            'Authorization', 'Content-Type', 'X-Device-Key',
            'X-Admin-Bootstrap-Token',
        ],
    )

# ==================== EPD 数据格式（7.3" E6，800x480，4bit a~p） ====================
EPD_WIDTH = 800
EPD_HEIGHT = 480
EPD_EXPECTED_CHARS = EPD_WIDTH * EPD_HEIGHT  # 384000
EPD_ALLOWED_CHARS = set('abcdefghijklmnop')
DEVICE_ID_PATTERN = re.compile(r'^(?:[0-9A-F]{6}|[0-9A-F]{12})$')
DEFAULT_SLEEP_INTERVAL_SECONDS = 12 * 60 * 60
MIN_SLEEP_INTERVAL_SECONDS = 5 * 60
MAX_SLEEP_INTERVAL_SECONDS = 30 * 24 * 60 * 60
ALLOW_REGISTRATION = os.environ.get('ALLOW_REGISTRATION', 'false').strip().lower() in ('1', 'true', 'yes', 'on')
AUTH_TOKEN_TTL_SECONDS = max(300, int(os.environ.get('AUTH_TOKEN_TTL_SECONDS', 7 * 24 * 60 * 60)))
DEVICE_AUTH_REQUIRED = Config.DEVICE_AUTH_REQUIRED
ADMIN_BOOTSTRAP_TOKEN = Config.ADMIN_BOOTSTRAP_TOKEN
PAIRING_MAX_FAILED_ATTEMPTS = Config.PAIRING_MAX_FAILED_ATTEMPTS
PAIRING_LOCK_SECONDS = Config.PAIRING_LOCK_SECONDS
DEVICE_STATUS_MAX_BODY_BYTES = Config.DEVICE_STATUS_MAX_BODY_BYTES
DEVICE_KEY_RESET_WINDOW_SECONDS = Config.DEVICE_KEY_RESET_WINDOW_SECONDS
UNCLAIMED_DEVICE_TTL_SECONDS = Config.UNCLAIMED_DEVICE_TTL_SECONDS
TELEMETRY_NUMERIC_LIMITS = {
    'rssi': (-150, 20),
    'uptime_ms': (0, 7 * 24 * 60 * 60 * 1000),
    'freeHeap': (0, 16 * 1024 * 1024),
    'currentSleepSeconds': (0, 366 * 24 * 60 * 60),
}


def sanitize_telemetry_text(value, limit: int) -> str:
    if not isinstance(value, str):
        return ''
    normalized = re.sub(r'[\x00-\x1f\x7f]+', ' ', value).strip()
    return normalized[:limit]


PAGE_MAX_NAME_CHARS = 100
PAGE_MAX_TYPE_CHARS = 32
PAGE_MAX_DATA_BYTES = 4 * 1024 * 1024
PAGE_MAX_THUMBNAIL_BYTES = 256 * 1024
NAMEPLATE_LOGO_MAX_BYTES = 512 * 1024
NAMEPLATE_LOGO_MAX_DIMENSION = 4096
NAMEPLATE_LOGO_MAX_PIXELS = 4096 * 4096
NAMEPLATE_LOGO_MIME_FORMATS = {
    'image/png': 'PNG',
    'image/jpeg': 'JPEG',
    'image/webp': 'WEBP',
}
NAMEPLATE_LOGO_CONFIG_KEYS = ('logoDataUrl', 'logoFileName', 'logoX', 'logoY')
TEMPLATE_DAY_TIMEZONE = os.environ.get('TEMPLATE_DAY_TIMEZONE', 'Asia/Shanghai')
try:
    TEMPLATE_DAY_TZ = ZoneInfo(TEMPLATE_DAY_TIMEZONE)
except Exception:
    TEMPLATE_DAY_TZ = timezone(timedelta(hours=8))
TEMPLATE_SLEEP_INTERVAL_SECONDS = {
    'weather': 6 * 60 * 60,
    'calendar': 24 * 60 * 60,
    'quote': 24 * 60 * 60,
    'todo': DEFAULT_SLEEP_INTERVAL_SECONDS,
    'qrcode': DEFAULT_SLEEP_INTERVAL_SECONDS,
    'nameplate': DEFAULT_SLEEP_INTERVAL_SECONDS,
}
CONTENT_MODE_LABELS = {
    'image': '普通图片',
    'text': '文字内容',
    'mixed': '图文内容',
    'custom': '自定义内容',
}


def utcnow() -> datetime:
    """Naive UTC for compatibility with existing PyMongo documents."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def normalize_sleep_interval(value, fallback: int) -> int:
    """Return a finite, firmware-safe sleep interval in seconds."""
    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and (isinstance(value, int) or math.isfinite(value))
        and value > 0
    ):
        return min(max(int(value), MIN_SLEEP_INTERVAL_SECONDS), MAX_SLEEP_INTERVAL_SECONDS)
    return min(max(int(fallback), MIN_SLEEP_INTERVAL_SECONDS), MAX_SLEEP_INTERVAL_SECONDS)

def get_template_meta(template_id: str):
    """查找当前仍可用的内置模板。"""
    if not template_id:
        return None
    template_id = str(template_id).strip().lower()
    return next((t for t in TEMPLATES if t.get('templateId') == template_id), None)

def build_content_metadata(content_mode=None, template_id=None, custom_sleep_interval=None):
    """统一计算设备当前内容标签和云端期望的唤醒间隔。

    custom_sleep_interval: 用户自定义的唤醒间隔（秒），优先于硬编码默认值。
    """
    mode = (content_mode or 'image')
    mode = str(mode).strip().lower() if isinstance(mode, str) else 'image'
    template_id = str(template_id).strip().lower() if template_id else None

    if mode == 'template':
        template = get_template_meta(template_id)
        if template:
            # 优先使用用户自定义间隔，否则用模板默认间隔
            sleep_interval = normalize_sleep_interval(
                custom_sleep_interval,
                TEMPLATE_SLEEP_INTERVAL_SECONDS.get(template_id, DEFAULT_SLEEP_INTERVAL_SECONDS),
            )
            return {
                'activeContentMode': 'template',
                'activeTemplateId': template_id,
                'activeContentLabel': f"{template.get('name', template_id)}模板",
                'sleepIntervalSeconds': sleep_interval,
            }
        mode = 'custom'

    if mode not in CONTENT_MODE_LABELS:
        mode = 'image'

    return {
        'activeContentMode': mode,
        'activeTemplateId': None,
        'activeContentLabel': CONTENT_MODE_LABELS[mode],
        'sleepIntervalSeconds': DEFAULT_SLEEP_INTERVAL_SECONDS,
    }

def get_device_content_metadata(device):
    if not device:
        return build_content_metadata()
    # 从 templateConfig 中读取用户自定义的唤醒间隔
    template_config = device.get('templateConfig', {})
    custom_interval = template_config.get('sleepIntervalSeconds') if isinstance(template_config, dict) else None
    meta = build_content_metadata(device.get('activeContentMode'), device.get('activeTemplateId'), custom_interval)
    if meta.get('activeTemplateId') == 'nameplate' and isinstance(template_config, dict):
        name = str(template_config.get('name') or template_config.get('personName') or '').strip()
        if name:
            meta['activeContentLabel'] = f'铭牌：{name}'
    return meta

def get_template_local_date(value):
    """Return the calendar date used by daily templates."""
    if not value:
        return None
    if hasattr(value, 'tzinfo'):
        dt = value
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(TEMPLATE_DAY_TZ).date()
    return None

def _should_re_render_template(template_id: str, config: dict, device: dict, wake_type: str = '') -> bool:
    """判断模板设备是否需要重渲染内容。

    按需渲染逻辑（设备唤醒时调用）：
    - 动态模板优先：天气每次渲染，每日一言手动唤醒强制刷新、定时唤醒按日期刷新
    - 二维码/待办：永不自动渲染
    - 其他 Canvas 发布内容：不自动覆盖用户设计
    """
    template_id = str(template_id).strip().lower() if template_id else ''
    wake_type = str(wake_type or '').strip().lower()
    render_source = device.get('renderSource')

    # 二维码/待办：内容不自动变化，永不自动渲染
    if template_id in ('qrcode', 'todo', 'nameplate'):
        print(f'🛑 _should_re_render: 模板={template_id}, renderSource={render_source} → 永不自动渲染')
        return False

    # 天气：每次唤醒都渲染（设备醒来就是要拿最新天气）
    if template_id == 'weather':
        print(f'🔄 _should_re_render: 模板={template_id}, renderSource={render_source} → 需要渲染（天气实时更新）')
        return True

    last_updated = device.get('activeContentUpdatedAt')
    today = datetime.now(TEMPLATE_DAY_TZ).date()

    # 每日一言：今天没渲染过 → 渲染
    if template_id == 'quote':
        if wake_type == 'manual':
            print(f'🔄 _should_re_render: 模板={template_id}, wakeType=manual → 需要渲染（手动唤醒）')
            return True
        if not last_updated:
            print(f'🔄 _should_re_render: 模板={template_id}, renderSource={render_source} → 需要渲染（无上次记录）')
            return True
        # 有上次更新时间：检查是否跨日
        last_date = get_template_local_date(last_updated)
        if last_date and last_date == today:
            print(f'🛑 _should_re_render: 模板={template_id}, renderSource={render_source} → 跳过（今天已渲染，{TEMPLATE_DAY_TIMEZONE}）')
            return False  # 今天已渲染过
        print(f'🔄 _should_re_render: 模板={template_id}, renderSource={render_source} → 需要渲染（跨日，{last_date} -> {today}）')
        return True

    # 日历：日期变了 → 渲染
    if template_id == 'calendar':
        if not last_updated:
            print(f'🔄 _should_re_render: 模板={template_id}, renderSource={render_source} → 需要渲染（无上次记录）')
            return True
        last_date = get_template_local_date(last_updated)
        if last_date and last_date == today:
            print(f'🛑 _should_re_render: 模板={template_id}, renderSource={render_source} → 跳过（今天已渲染，{TEMPLATE_DAY_TIMEZONE}）')
            return False  # 今天已渲染过（日期没变）
        print(f'🔄 _should_re_render: 模板={template_id}, renderSource={render_source} → 需要渲染（跨日，{last_date} -> {today}）')
        return True

    # 其他 Canvas 发布内容：不自动覆盖，保留用户设计的效果
    if render_source == 'canvas':
        print(f'🛑 _should_re_render: 模板={template_id}, renderSource=canvas → 跳过（保留用户设计）')
        return False

    # 其他未知模板：保守策略，不自动渲染
    print(f'🛑 _should_re_render: 模板={template_id}, renderSource={render_source} → 跳过（未知模板）')
    return False


def render_template_epd_data(template_id: str, config: dict) -> str:
    """Render template through the full preview/data pipeline and return EPD data only."""
    result = render_template_with_preview(template_id, config)
    epd_data = result.get('epdData') if isinstance(result, dict) else None
    if not epd_data or len(epd_data) != EPD_EXPECTED_CHARS:
        raise ValueError(f'Template rendering failed or invalid length: {len(epd_data) if epd_data else 0}')
    return epd_data


NAMEPLATE_MAX_NAME_LEN = 16
NAMEPLATE_MAX_NAMES = 500
NAMEPLATE_MAX_TARGET_DEVICES = 500
NAMEPLATE_MAX_PARSE_FILE_BYTES = int(os.environ.get('NAMEPLATE_MAX_PARSE_FILE_BYTES', 8 * 1024 * 1024))
NAMEPLATE_MAX_PARSE_FILES = max(1, int(os.environ.get('NAMEPLATE_MAX_PARSE_FILES', 8)))
NAMEPLATE_MAX_PARSE_TOTAL_BYTES = int(os.environ.get('NAMEPLATE_MAX_PARSE_TOTAL_BYTES', 16 * 1024 * 1024))
NAMEPLATE_MAX_PARSE_TEXT_CHARS = int(os.environ.get('NAMEPLATE_MAX_PARSE_TEXT_CHARS', 20_000))
NAMEPLATE_MAX_XLSX_ENTRIES = 200
NAMEPLATE_MAX_XLSX_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
# Keep upstream AI parsing and serial batch rendering below Gunicorn's 120s
# worker timeout. The request helpers consume these as end-to-end deadlines,
# rather than granting every retry a fresh timeout.
NAMEPLATE_AI_TOTAL_BUDGET_SECONDS = 90.0
NAMEPLATE_DISPATCH_DEADLINE_SECONDS = 90.0
NAMEPLATE_AI_API_KEY = (
    os.environ.get('NAMEPLATE_AI_API_KEY')
    or os.environ.get('OPENAI_API_KEY')
    or ''
).strip()
NAMEPLATE_AI_BASE_URL = (
    os.environ.get('NAMEPLATE_AI_BASE_URL')
    or os.environ.get('OPENAI_BASE_URL')
    or 'https://api.openai.com/v1'
).rstrip('/')
NAMEPLATE_AI_MODEL = (
    os.environ.get('NAMEPLATE_AI_MODEL')
    or os.environ.get('OPENAI_NAMEPLATE_MODEL')
    or 'gpt-4.1-mini'
)
NAMEPLATE_AI_API_MODE = os.environ.get('NAMEPLATE_AI_API_MODE', 'responses').strip().lower()
NAMEPLATE_LABEL_PREFIXES = {
    '名单', '姓名', '人员', '参会人', '参会人员', '嘉宾', '领导', '铭牌', '下发名单', '姓名名单'
}
NAMEPLATE_REJECT_VALUES = {
    '姓名', '名字', '名单', '人员', '参会人', '参会人员', '嘉宾', '领导', '单位',
    '部门', '职务', '职位', '序号', '编号', '备注', '电话', '手机', '设备', '设备号',
    'device', 'name', 'title', 'department', 'position', 'no'
}
NAMEPLATE_LEADING_WORDS = (
    '请给', '请为', '请把', '请将', '请', '给', '为', '把', '将',
    '下发', '发送', '显示', '设置', '安排', '名单', '姓名', '人员', '铭牌'
)


def get_nameplate_ai_api_key() -> str:
    return (
        os.environ.get('NAMEPLATE_AI_API_KEY')
        or os.environ.get('OPENAI_API_KEY')
        or NAMEPLATE_AI_API_KEY
        or ''
    ).strip()


def _clean_nameplate_candidate(raw_value) -> str:
    if raw_value is None:
        return ''

    value = str(raw_value).strip()
    if not value:
        return ''

    value = re.sub(r'^[\s\-\*\u2022]+', '', value)
    value = re.sub(r'^\d+[\.\、\)\）\s]+', '', value)
    value = re.sub(r'^[一二三四五六七八九十百]+[\.\、\)\）\s]+', '', value)
    value = value.strip(' "\'“”‘’[]【】()（）<>《》')

    if ':' in value or '：' in value:
        label, rest = re.split(r'[:：]', value, maxsplit=1)
        if label.strip() in NAMEPLATE_LABEL_PREFIXES:
            value = rest.strip()

    changed = True
    while changed and value:
        changed = False
        for word in NAMEPLATE_LEADING_WORDS:
            if value.startswith(word):
                value = value[len(word):].strip()
                changed = True

    value = re.split(r'[（(【\[]', value, maxsplit=1)[0].strip()
    value = re.split(r'[-—/|]', value, maxsplit=1)[0].strip()
    if re.search(r'\s', value):
        value = re.split(r'\s+', value, maxsplit=1)[0].strip()

    value = re.sub(r'(同志|先生|女士|老师)$', '', value).strip()
    value = value.strip(' "\'“”‘’.,，。;；:：!?！？[]【】()（）<>《》')

    if value.lower() in NAMEPLATE_REJECT_VALUES or value in NAMEPLATE_REJECT_VALUES:
        return ''
    if not value or len(value) > NAMEPLATE_MAX_NAME_LEN:
        return ''
    if not re.search(r'[\u4e00-\u9fffA-Za-z]', value):
        return ''
    if re.search(r'\d', value):
        return ''
    if re.search(r'[，,。；;:：!?！？]', value):
        return ''
    return value


def parse_nameplate_names(data: dict) -> list[str]:
    """Parse a conservative name list from API payload.

    Future AI/WeChat integrations should call this same dispatch API with a
    normalized names array after completing intent extraction.
    """
    names = []

    def append_name(raw_value):
        name = _clean_nameplate_candidate(raw_value)
        if name and len(names) <= NAMEPLATE_MAX_NAMES:
            names.append(name)
    raw_names = data.get('names')
    if isinstance(raw_names, list):
        for item in raw_names:
            append_name(item)
            if len(names) > NAMEPLATE_MAX_NAMES:
                break
        return names

    text = data.get('text') or data.get('message') or ''
    if not isinstance(text, str):
        return []

    normalized = text.replace('\r\n', '\n').replace('\r', '\n')
    normalized = re.sub(r'[、,，;；\t]+', '\n', normalized)

    for line in normalized.split('\n'):
        append_name(line)
        if len(names) > NAMEPLATE_MAX_NAMES:
            break

    return names


def parse_nameplate_names_from_text(text: str) -> list[str]:
    if not isinstance(text, str) or not text.strip():
        return []

    names = []
    normalized = text.replace('\r\n', '\n').replace('\r', '\n')
    for line in normalized.split('\n'):
        cells = re.split(r'[\t,，、;；|]+', line)
        for cell in cells:
            name = _clean_nameplate_candidate(cell)
            if name:
                names.append(name)
                break
        if len(names) > NAMEPLATE_MAX_NAMES:
            break
    return names


def resolve_nameplate_target_devices(owner: str, requested_device_ids) -> tuple[list[dict], list[str]]:
    if devices_collection is None or not owner:
        return [], []

    if isinstance(requested_device_ids, list) and requested_device_ids:
        clean_ids = []
        for raw_id in requested_device_ids:
            clean_id = normalize_device_id(str(raw_id))
            if clean_id and clean_id not in clean_ids:
                clean_ids.append(clean_id)
            if len(clean_ids) > NAMEPLATE_MAX_TARGET_DEVICES:
                break

        docs = list(devices_collection.find({'owner': owner, 'deviceId': {'$in': clean_ids}}))
        by_id = {doc.get('deviceId'): doc for doc in docs}
        target_devices = [by_id[device_id] for device_id in clean_ids if device_id in by_id]
        missing = [device_id for device_id in clean_ids if device_id not in by_id]
        return target_devices, missing

    return list(
        devices_collection.find({'owner': owner, 'claimed': True})
        .sort('addedAt', 1)
        .limit(NAMEPLATE_MAX_TARGET_DEVICES)
    ), []


def normalize_nameplate_logo_data_url(value) -> str:
    """Validate and canonicalize a user-supplied raster logo data URL."""
    if not isinstance(value, str) or not value.startswith('data:image/'):
        return ''

    header, separator, payload = value.partition(',')
    if not separator or not header.endswith(';base64'):
        return ''
    mime_type = header[5:-7].lower()
    expected_format = NAMEPLATE_LOGO_MIME_FORMATS.get(mime_type)
    if not expected_format or not payload:
        return ''

    try:
        raw = base64.b64decode(payload, validate=True)
    except Exception:
        return ''
    if not raw or len(raw) > NAMEPLATE_LOGO_MAX_BYTES:
        return ''

    try:
        with Image.open(io.BytesIO(raw)) as logo:
            if (logo.format or '').upper() != expected_format:
                return ''
            width, height = logo.size
            if (
                width <= 0 or height <= 0
                or width > NAMEPLATE_LOGO_MAX_DIMENSION
                or height > NAMEPLATE_LOGO_MAX_DIMENSION
                or width * height > NAMEPLATE_LOGO_MAX_PIXELS
            ):
                return ''
            logo.verify()
    except Exception:
        return ''

    encoded = base64.b64encode(raw).decode('ascii')
    return f'data:{mime_type};base64,{encoded}'


def merge_nameplate_logo_config(config: dict, source_config: dict) -> dict:
    """Keep logo payload and placement out of AI edits while preserving the template design."""
    merged = dict(config or {})
    source = source_config if isinstance(source_config, dict) else {}
    for key in NAMEPLATE_LOGO_CONFIG_KEYS:
        if key in source:
            merged[key] = source[key]
    return merged


def normalize_nameplate_template_config(raw_config) -> dict:
    raw_config = raw_config if isinstance(raw_config, dict) else {}
    style = str(raw_config.get('backgroundStyle') or 'formal_red').strip().lower()
    if style not in ('formal_red', 'formal_blue', 'formal_green', 'plain'):
        style = 'formal_red'

    config = {
        'backgroundStyle': style,
        'title': str(raw_config.get('title') or '').strip()[:40],
        'subtitle': str(raw_config.get('subtitle') or '').strip()[:40],
    }

    logo_data_url = normalize_nameplate_logo_data_url(raw_config.get('logoDataUrl'))
    if logo_data_url:
        config['logoDataUrl'] = logo_data_url
        logo_file_name = str(raw_config.get('logoFileName') or '').strip()
        if logo_file_name:
            config['logoFileName'] = logo_file_name[:100]

    logo_x = raw_config.get('logoX')
    logo_y = raw_config.get('logoY')
    if (
        isinstance(logo_x, (int, float)) and not isinstance(logo_x, bool)
        and isinstance(logo_y, (int, float)) and not isinstance(logo_y, bool)
        and math.isfinite(logo_x) and math.isfinite(logo_y)
    ):
        config['logoX'] = min(max(int(round(logo_x)), 0), EPD_WIDTH - 1)
        config['logoY'] = min(max(int(round(logo_y)), 0), EPD_HEIGHT - 1)

    sleep_interval = raw_config.get('sleepIntervalSeconds')
    if isinstance(sleep_interval, (int, float)) and not isinstance(sleep_interval, bool):
        config['sleepIntervalSeconds'] = normalize_sleep_interval(
            sleep_interval, DEFAULT_SLEEP_INTERVAL_SECONDS
        )
    return config


def normalize_nameplate_template_name(value) -> str:
    name = str(value or '').strip()
    return (name[:40] if name else '会议名牌模板')


def serialize_saved_nameplate_template(doc):
    if not doc:
        return None

    result = {
        'templateId': doc.get('templateId'),
        'name': doc.get('name') or '会议名牌模板',
        'baseTemplateId': doc.get('baseTemplateId') or 'nameplate',
        'templateConfig': normalize_nameplate_template_config(doc.get('templateConfig')),
    }
    for key in ('createdAt', 'updatedAt'):
        value = doc.get(key)
        if hasattr(value, 'isoformat'):
            result[key] = value.isoformat()
        elif value:
            result[key] = str(value)
    return result


def find_page_for_owner(page_id: str, owner: str, projection=None):
    """Prefer owner-scoped pages, then fall back to one legacy ownerless page."""
    if pages_collection is None or not owner:
        return None
    page = pages_collection.find_one(
        {'pageId': page_id, 'owner': owner}, projection,
    )
    if page:
        return page
    return pages_collection.find_one(
        {'pageId': page_id, 'owner': {'$exists': False}}, projection,
    )


def decode_uploaded_text(raw: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-8', 'gb18030', 'latin-1'):
        try:
            return raw.decode(encoding)
        except Exception:
            continue
    return raw.decode('utf-8', errors='ignore')


def validate_xlsx_archive(raw: bytes) -> None:
    """Reject malformed or excessively expanded XLSX/ZIP containers before parsing."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            entries = archive.infolist()
            if len(entries) > NAMEPLATE_MAX_XLSX_ENTRIES:
                raise ValueError('表格压缩包文件项过多')
            total_uncompressed = 0
            for entry in entries:
                if entry.is_dir():
                    continue
                if entry.file_size < 0 or entry.compress_size < 0:
                    raise ValueError('表格压缩包元数据无效')
                total_uncompressed += entry.file_size
                if total_uncompressed > NAMEPLATE_MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError('表格解压后体积超过限制')
    except zipfile.BadZipFile as exc:
        raise ValueError('表格不是有效的 XLSX 压缩包') from exc


def extract_spreadsheet_text(raw: bytes, filename: str) -> tuple[str, list[str]]:
    warnings = []
    suffix = Path(filename or '').suffix.lower()
    if suffix in ('.csv', '.tsv', '.txt'):
        return decode_uploaded_text(raw), warnings

    if suffix in ('.xlsx', '.xlsm'):
        try:
            validate_xlsx_archive(raw)
        except ValueError as exc:
            return '', [str(exc)]

        try:
            from openpyxl import load_workbook
        except Exception:
            return '', ['服务器缺少 openpyxl，暂不能解析 XLSX 表格']

        try:
            wb = load_workbook(
                io.BytesIO(raw), read_only=True, data_only=True, keep_links=False,
            )
            lines = []
            for sheet in wb.worksheets[:5]:
                lines.append(f'工作表: {sheet.title}')
                for row in sheet.iter_rows(max_row=300, max_col=30, values_only=True):
                    cells = []
                    for cell in row:
                        if cell is None:
                            continue
                        text = str(cell).strip()
                        if text:
                            cells.append(text)
                    if cells:
                        lines.append('\t'.join(cells))
            return '\n'.join(lines), warnings
        except Exception as e:
            print(f'❌ 表格解析失败: {e}')
            return '', ['表格解析失败']

    return '', [f'不支持的表格格式: {suffix or filename}']


def collect_nameplate_parse_sources(req) -> tuple[str, list[dict], list[str], list[str]]:
    warnings = []
    filenames = []
    text_parts = []
    image_parts = []

    if req.content_type and req.content_type.startswith('application/json'):
        data = req.get_json() or {}
        text = data.get('text') or data.get('message') or ''
        if isinstance(text, str) and text.strip():
            if len(text) > NAMEPLATE_MAX_PARSE_TEXT_CHARS:
                raise ValueError(f'文本长度不能超过 {NAMEPLATE_MAX_PARSE_TEXT_CHARS} 字符')
            text_parts.append(text)
        return '\n'.join(text_parts), image_parts, warnings, filenames

    form_text = req.form.get('text') or req.form.get('message') or ''
    if form_text.strip():
        if len(form_text) > NAMEPLATE_MAX_PARSE_TEXT_CHARS:
            raise ValueError(f'文本长度不能超过 {NAMEPLATE_MAX_PARSE_TEXT_CHARS} 字符')
        text_parts.append(form_text.strip())

    uploads = req.files.getlist('files')
    if len(uploads) > NAMEPLATE_MAX_PARSE_FILES:
        raise ValueError(f'上传文件数不能超过 {NAMEPLATE_MAX_PARSE_FILES} 个')

    total_bytes = 0
    for storage in uploads:
        filename = storage.filename or 'upload'
        raw = storage.read(NAMEPLATE_MAX_PARSE_FILE_BYTES + 1)
        if not raw:
            continue
        filenames.append(filename)
        if len(raw) > NAMEPLATE_MAX_PARSE_FILE_BYTES:
            raise ValueError(f'{filename} 超过单文件大小限制')
        total_bytes += len(raw)
        if total_bytes > NAMEPLATE_MAX_PARSE_TOTAL_BYTES:
            raise ValueError('上传文件总大小超过限制')

        mime_type = storage.mimetype or mimetypes.guess_type(filename)[0] or ''
        suffix = Path(filename).suffix.lower()
        if mime_type.startswith('image/') or suffix in ('.png', '.jpg', '.jpeg', '.webp', '.gif'):
            if suffix == '.gif':
                warnings.append(f'{filename} 如为动图，仅建议上传静态图片')
            image_parts.append({
                'filename': filename,
                'mimeType': mime_type or 'image/png',
                'dataUrl': f"data:{mime_type or 'image/png'};base64,{base64.b64encode(raw).decode('ascii')}",
            })
            continue

        extracted_text, file_warnings = extract_spreadsheet_text(raw, filename)
        warnings.extend(file_warnings)
        if extracted_text:
            text_parts.append(f'文件: {filename}\n{extracted_text}')

    combined_text = '\n\n'.join(text_parts)
    if len(combined_text) > NAMEPLATE_MAX_PARSE_TEXT_CHARS:
        raise ValueError(f'解析后文本长度不能超过 {NAMEPLATE_MAX_PARSE_TEXT_CHARS} 字符')
    return combined_text, image_parts, warnings, filenames


def build_nameplate_parse_result(names: list[str], template_config: dict, warnings=None,
                                 ai_used=False, source_summary='') -> dict:
    clean_names = []
    for name in names:
        clean = _clean_nameplate_candidate(name)
        if clean:
            clean_names.append(clean)
        if len(clean_names) >= NAMEPLATE_MAX_NAMES:
            break

    result_warnings = list(warnings or [])
    if len(names) > len(clean_names) and len(clean_names) >= NAMEPLATE_MAX_NAMES:
        result_warnings.append(f'名单最多保留前 {NAMEPLATE_MAX_NAMES} 个姓名')

    return {
        'names': clean_names,
        'templateConfig': normalize_nameplate_template_config(template_config),
        'warnings': result_warnings,
        'aiUsed': bool(ai_used),
        'sourceSummary': source_summary or '',
    }


def _extract_openai_response_text(resp_json: dict) -> str:
    if not isinstance(resp_json, dict):
        return ''
    if isinstance(resp_json.get('output_text'), str):
        return resp_json['output_text']

    texts = []
    for output in resp_json.get('output', []) or []:
        for content in output.get('content', []) or []:
            if isinstance(content.get('text'), str):
                texts.append(content['text'])
    return '\n'.join(texts)


def _nameplate_ai_request_timeout(deadline: float) -> tuple[float, float]:
    """Return connect/read timeouts whose combined budget fits the deadline."""
    remaining = deadline - time.monotonic()
    if remaining <= 2.0:
        raise TimeoutError('AI解析时间预算已用尽')
    connect_timeout = min(10.0, max(0.5, remaining * 0.2))
    read_timeout = remaining - connect_timeout
    return connect_timeout, read_timeout


def call_openai_nameplate_parser(source_text: str, image_parts: list[dict], base_config: dict) -> dict:
    api_key = get_nameplate_ai_api_key()
    if not api_key:
        raise RuntimeError('NAMEPLATE_AI_API_KEY 未配置')
    request_deadline = time.monotonic() + NAMEPLATE_AI_TOTAL_BUDGET_SECONDS

    schema = {
        'type': 'object',
        'additionalProperties': False,
        'required': ['names', 'templateConfig', 'warnings', 'sourceSummary'],
        'properties': {
            'names': {
                'type': 'array',
                'items': {'type': 'string'},
                'description': '按原始顺序提取的人名，只保留姓名，不要包含单位、职务、序号。'
            },
            'templateConfig': {
                'type': 'object',
                'additionalProperties': False,
                'required': ['backgroundStyle', 'title', 'subtitle', 'sleepIntervalSeconds'],
                'properties': {
                    'backgroundStyle': {'type': 'string', 'enum': ['formal_red', 'formal_blue', 'formal_green', 'plain']},
                    'title': {'type': 'string'},
                    'subtitle': {'type': 'string'},
                    'sleepIntervalSeconds': {'type': 'integer'},
                },
            },
            'warnings': {'type': 'array', 'items': {'type': 'string'}},
            'sourceSummary': {'type': 'string'},
        },
    }

    prompt_config = {
        key: value for key, value in normalize_nameplate_template_config(base_config).items()
        if key not in NAMEPLATE_LOGO_CONFIG_KEYS
    }
    prompt_text = (
        '你是政务会议电子铭牌名单解析助手。请从用户上传的文字、图片或表格中提取需要下发到铭牌的姓名。'
        '输出必须符合 JSON Schema。只提取人名，不要把单位、职务、标题、设备编号、电话、序号当作姓名。'
        '保持名单原始顺序。若文本中出现职务或英文副标题，可作为 title；公司名称可作为 subtitle。'
        'backgroundStyle 可使用 formal_red=Pheno红色底栏、formal_green=Pheno绿色底栏、plain=Pheno绿色横幅、formal_blue=Pheno职务名片。'
        '如果不确定，请把疑问写入 warnings。'
        f'\n当前默认模板: {json.dumps(prompt_config, ensure_ascii=False)}'
        f'\n文本内容:\n{source_text[:12000]}'
    )

    base_url = (
        os.environ.get('NAMEPLATE_AI_BASE_URL')
        or os.environ.get('OPENAI_BASE_URL')
        or NAMEPLATE_AI_BASE_URL
    ).rstrip('/')
    model = (
        os.environ.get('NAMEPLATE_AI_MODEL')
        or os.environ.get('OPENAI_NAMEPLATE_MODEL')
        or NAMEPLATE_AI_MODEL
    )
    api_mode = (os.environ.get('NAMEPLATE_AI_API_MODE') or NAMEPLATE_AI_API_MODE or 'responses').strip().lower()

    if api_mode in ('chat', 'chat_completions', 'chat-completions'):
        content = [{'type': 'text', 'text': prompt_text}]
        for image in image_parts[:8]:
            content.append({
                'type': 'image_url',
                'image_url': {
                    'url': image['dataUrl'],
                    'detail': 'high',
                },
            })

        payload = {
            'model': model,
            'messages': [{'role': 'user', 'content': content}],
            'response_format': {'type': 'json_object'},
            'max_tokens': 1800,
        }
        if 'minimax' in base_url.lower() or model.lower().startswith('minimax-'):
            payload['thinking'] = {'type': 'disabled'}

        resp = requests.post(
            f'{base_url}/chat/completions',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            json=payload,
            timeout=_nameplate_ai_request_timeout(request_deadline),
        )
        if resp.status_code >= 400 and 'response_format' in resp.text:
            retry_payload = dict(payload)
            retry_payload.pop('response_format', None)
            resp = requests.post(
                f'{base_url}/chat/completions',
                headers={
                    'Authorization': f'Bearer {api_key}',
                    'Content-Type': 'application/json',
                },
                json=retry_payload,
                timeout=_nameplate_ai_request_timeout(request_deadline),
            )
        if resp.status_code >= 400:
            raise RuntimeError(f'AI解析失败: HTTP {resp.status_code} {resp.text[:300]}')

        completion = resp.json()
        output_text = (
            completion.get('choices', [{}])[0]
            .get('message', {})
            .get('content', '')
        )
        if not output_text:
            raise RuntimeError('AI解析未返回文本')

        parsed = json.loads(output_text)
        parsed_config = merge_nameplate_logo_config(
            parsed.get('templateConfig', base_config), base_config
        )
        return build_nameplate_parse_result(
            parsed.get('names', []),
            parsed_config,
            warnings=parsed.get('warnings', []),
            ai_used=True,
            source_summary=parsed.get('sourceSummary', 'AI解析'),
        )

    content = [{
        'type': 'input_text',
        'text': prompt_text
    }]

    for image in image_parts[:8]:
        content.append({
            'type': 'input_image',
            'image_url': image['dataUrl'],
            'detail': 'high',
        })

    payload = {
        'model': model,
        'input': [{'role': 'user', 'content': content}],
        'text': {
            'format': {
                'type': 'json_schema',
                'name': 'nameplate_parse_result',
                'strict': True,
                'schema': schema,
            }
        },
        'max_output_tokens': 1800,
    }

    resp = requests.post(
        f'{base_url}/responses',
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        json=payload,
        timeout=_nameplate_ai_request_timeout(request_deadline),
    )
    if resp.status_code >= 400:
        raise RuntimeError(f'OpenAI 解析失败: HTTP {resp.status_code} {resp.text[:300]}')

    output_text = _extract_openai_response_text(resp.json())
    if not output_text:
        raise RuntimeError('OpenAI 解析未返回文本')
    try:
        parsed = json.loads(output_text)
    except Exception as e:
        raise RuntimeError(f'OpenAI 解析结果不是 JSON: {e}')

    parsed_config = merge_nameplate_logo_config(
        parsed.get('templateConfig') or base_config, base_config
    )
    return build_nameplate_parse_result(
        parsed.get('names', []),
        parsed_config,
        parsed.get('warnings') or [],
        ai_used=True,
        source_summary=parsed.get('sourceSummary') or '',
    )


def to_epoch_ms(value):
    if not value:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if hasattr(value, 'timestamp'):
        return int(value.timestamp() * 1000)
    return None

def validate_epd_text_payload(image_data: str):
    """校验 EPD 原始数据（a~p 编码字符串）是否完整且合法。

    目标：尽量把“数据不完整/格式异常”的问题拦在云端发布阶段，
    避免设备端下载后才发现缺数据导致白屏。
    """
    if not isinstance(image_data, str) or not image_data:
        return False, 'Empty image data'
    if len(image_data) != EPD_EXPECTED_CHARS:
        return False, f'Invalid length: expected {EPD_EXPECTED_CHARS}, got {len(image_data)}'
    # 快速字符集校验（a~p），允许集合最多 16 种字符，set() 成本很小
    invalid = set(image_data) - EPD_ALLOWED_CHARS
    if invalid:
        bad = ''.join(sorted(list(invalid))[:8])
        return False, f'Invalid chars: {bad}'
    return True, None

# ==================== MongoDB 连接 ====================
mongo_client = None
db = None
users_collection = None
devices_collection = None
device_status_collection = None
pages_collection = None
pairing_codes_collection = None
saved_nameplate_templates_collection = None

# ==================== 图片持久化存储目录 ====================
# 图片数据保存在 data/epd/<deviceId>/latest.txt
DATA_DIR = Path(__file__).parent / 'data' / 'epd'
DATA_DIR.mkdir(parents=True, exist_ok=True)
DEVICE_WRITE_LOCK_STRIPES = 128
_device_write_locks = tuple(threading.Lock() for _ in range(DEVICE_WRITE_LOCK_STRIPES))

def is_valid_device_id(device_id: str) -> bool:
    return bool(DEVICE_ID_PATTERN.fullmatch(normalize_device_id(device_id)))


def get_device_write_lock(device_id: str):
    clean_id = normalize_device_id(device_id)
    if not is_valid_device_id(clean_id):
        raise ValueError('Invalid deviceId format')
    # A bounded striped pool preserves same-device serialization without an
    # attacker being able to grow one permanent Lock object per arbitrary ID.
    return _device_write_locks[int(clean_id, 16) % DEVICE_WRITE_LOCK_STRIPES]


def get_device_data_dir(device_id: str, create: bool = False) -> Path:
    """获取设备数据目录"""
    clean_id = normalize_device_id(device_id)
    if not is_valid_device_id(clean_id):
        raise ValueError('Invalid deviceId format')
    device_dir = DATA_DIR / clean_id
    if create:
        device_dir.mkdir(parents=True, exist_ok=True)
    return device_dir

def get_device_image_path(device_id: str, create_parent: bool = False) -> Path:
    """获取设备最新图片文件路径"""
    return get_device_data_dir(device_id, create=create_parent) / 'latest.txt'


def get_ready_device_image_path(device_id: str, device=None):
    image_path = get_device_image_path(device_id, create_parent=False)
    try:
        if not image_path.is_file() or image_path.stat().st_size != EPD_EXPECTED_CHARS:
            return None
        declared_hash = None
        if device:
            declared_size = device.get('imageSizeChars')
            if declared_size is not None and declared_size != EPD_EXPECTED_CHARS:
                return None
            declared_hash = device.get('imageSha256')
            if declared_hash is not None:
                declared_hash = str(declared_hash).lower()
                if not re.fullmatch(r'[0-9a-f]{64}', declared_hash):
                    return None

        digest = hashlib.sha256() if declared_hash is not None else None
        actual_size = 0
        with image_path.open('rb') as image_file:
            for chunk in iter(lambda: image_file.read(64 * 1024), b''):
                actual_size += len(chunk)
                if any(value < ord('a') or value > ord('p') for value in chunk):
                    return None
                if digest is not None:
                    digest.update(chunk)
        if actual_size != EPD_EXPECTED_CHARS:
            return None
        if digest is not None and not secrets.compare_digest(digest.hexdigest(), declared_hash):
            return None
    except OSError:
        return None
    return image_path


def build_raw_image_url(device_id: str, image_version: int) -> str:
    clean_id = normalize_device_id(device_id)
    if not is_valid_device_id(clean_id):
        raise ValueError('Invalid deviceId format')
    return f'{Config.PUBLIC_BASE_URL}/api/epd/raw/{clean_id}?v={int(image_version)}'


def cleanup_device_artifacts(device_id: str) -> None:
    """Remove tenant-scoped pages, pairing state, telemetry and image files."""
    clean_id = normalize_device_id(device_id)
    if not is_valid_device_id(clean_id):
        raise ValueError('Invalid deviceId format')

    page_device_ids = device_id_variants(clean_id)
    if pages_collection is not None:
        pages_collection.delete_many({'deviceId': {'$in': page_device_ids}})
    if pairing_codes_collection is not None:
        pairing_codes_collection.delete_one({'deviceId': clean_id})
    if device_status_collection is not None:
        # 保留 deviceKeyHash，避免解绑后攻击者重新 TOFU 抢占设备身份。
        device_status_collection.update_one(
            {'deviceId': clean_id},
            {'$unset': {
                'lastSeen': '', 'updatedAt': '', 'remoteIp': '', 'ip': '',
                'rssi': '', 'uptime_ms': '', 'freeHeap': '',
                'currentSleepSeconds': '', 'lastWakeType': '',
                'lastWakeCause': '', 'lastManualWake': '', 'lastAutoWake': '',
                'deviceKeyResetUntil': '', 'deviceKeyResetRequestedBy': '',
                'deviceKeyResetRequestedAt': '',
                'unclaimedExpiresAt': '',
            }},
        )

    device_dir = get_device_data_dir(clean_id, create=False)
    if device_dir.exists():
        shutil.rmtree(device_dir)

def save_device_image(device_id: str, image_data: str) -> bool:
    """保存设备图片数据到磁盘"""
    tmp_path = None
    try:
        image_path = get_device_image_path(device_id, create_parent=True)
        fd, tmp_name = tempfile.mkstemp(prefix='.latest-', suffix='.tmp', dir=str(image_path.parent))
        tmp_path = Path(tmp_name)
        with os.fdopen(fd, 'w', encoding='utf-8', newline='') as f:
            f.write(image_data)
            f.flush()
            try:
                os.fsync(f.fileno())
            except Exception:
                # 某些环境/文件系统可能不支持 fsync，忽略但仍然 replace
                pass
        os.replace(tmp_path, image_path)
        tmp_path = None

        print(f'💾 图片已保存: {image_path} ({len(image_data)} 字符)')
        return True
    except Exception as e:
        print(f'❌ 保存图片失败: {e}')
        return False
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass

def load_device_image(device_id: str) -> str:
    """从磁盘加载设备图片数据"""
    try:
        image_path = get_device_image_path(device_id, create_parent=False)
        if image_path.exists():
            with open(image_path, 'r', encoding='utf-8') as f:
                return f.read()
    except Exception as e:
        print(f'❌ 加载图片失败: {e}')
    return None

def redact_uri_secret(uri: str) -> str:
    """Hide password-like credentials before printing a URI to logs."""
    try:
        parsed = urlsplit(uri)
        if parsed.password is None:
            return uri
        username = parsed.username or ''
        host = parsed.hostname or ''
        port = f':{parsed.port}' if parsed.port else ''
        netloc = f'{username}:***@{host}{port}'
        return urlunsplit((parsed.scheme, netloc, parsed.path, parsed.query, parsed.fragment))
    except Exception:
        return '<redacted-uri>'

def connect_mongodb(max_retries: int = 10, retry_delay_seconds: int = 2):
    """连接 MongoDB"""
    global mongo_client, db, users_collection, devices_collection, device_status_collection
    global pages_collection, pairing_codes_collection, saved_nameplate_templates_collection
    for attempt in range(1, max_retries + 1):
        try:
            mongo_client = MongoClient(Config.MONGODB_URI, serverSelectionTimeoutMS=5000)
            # 测试连接
            mongo_client.server_info()
            db = mongo_client[Config.MONGODB_DB]
            users_collection = db['users']
            devices_collection = db['devices']
            device_status_collection = db['device_status']
            pages_collection = db['pages']
            pairing_codes_collection = db['pairing_codes']
            saved_nameplate_templates_collection = db['saved_nameplate_templates']

            ensure_all_indexes(db)
            # 旧版明文 token 不再被接受，启动时主动失效。
            users_collection.update_many(
                {'token': {'$exists': True}},
                {'$unset': {'token': ''}},
            )

            print(f'✅ Connected to MongoDB: {redact_uri_secret(Config.MONGODB_URI)}')
            print(f'📊 Database: {Config.MONGODB_DB}')
            return True
        except Exception as e:
            if mongo_client is not None:
                try:
                    mongo_client.close()
                except Exception:
                    pass
            mongo_client = None
            db = None
            users_collection = None
            devices_collection = None
            device_status_collection = None
            pages_collection = None
            pairing_codes_collection = None
            saved_nameplate_templates_collection = None
            if attempt < max_retries:
                print(f'⚠️  MongoDB connection attempt {attempt}/{max_retries} failed: {e}')
                time.sleep(retry_delay_seconds)
                continue
            print(f'❌ MongoDB connection error after {max_retries} attempts: {e}')
            raise RuntimeError('MongoDB initialization failed') from e

# ==================== 用户认证工具函数 ====================

def hash_password(password: str) -> str:
    return generate_password_hash(password, method='scrypt')


DUMMY_PASSWORD_HASH = generate_password_hash('invalid-login-timing-sentinel', method='scrypt')


def verify_password(password_hash: str, password: str) -> tuple[bool, bool]:
    """Return (valid, needs_migration) for scrypt and legacy SHA-256 hashes."""
    if not password_hash:
        return False, False
    if password_hash.startswith('scrypt:'):
        try:
            return check_password_hash(password_hash, password), False
        except (TypeError, ValueError):
            return False, False

    legacy_hash = hashlib.sha256(password.encode('utf-8')).hexdigest()
    valid = len(password_hash) == 64 and secrets.compare_digest(password_hash.lower(), legacy_hash)
    return valid, valid

def generate_token() -> str:
    return secrets.token_hex(32)


def hash_token(token: str) -> str:
    return hashlib.sha256(token.encode('utf-8')).hexdigest()

def get_current_user():
    """根据 Authorization: Bearer <token> 获取当前用户"""
    global users_collection
    if users_collection is None:
        return None

    auth_header = request.headers.get('Authorization', '')
    if not auth_header.startswith('Bearer '):
        return None
    token = auth_header[7:].strip()
    if not token:
        return None

    user = users_collection.find_one({
        'tokenHash': hash_token(token),
        'tokenExpiresAt': {'$gt': utcnow()},
    })
    return user

def login_required(f):
    """需要登录的装饰器"""
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        request.user = user
        return f(*args, **kwargs)
    return wrapper

def normalize_device_id(device_id: str) -> str:
    """统一规范化 deviceId：去掉分隔符并转大写。

    设备端/前端可能传入带 '-' ':' 或小写的 ID；数据库里统一存 clean uppercase。
    """
    if not device_id:
        return ''
    return (device_id or '').strip().upper().replace('-', '').replace(':', '')


def device_id_variants(device_id: str) -> list[str]:
    """Return normalized and legacy separator/case variants for page migration."""
    raw_id = str(device_id or '').strip()
    clean_id = normalize_device_id(raw_id)
    variants = {raw_id, raw_id.upper(), raw_id.lower(), clean_id, clean_id.lower()}
    if is_valid_device_id(clean_id):
        pairs = [clean_id[index:index + 2] for index in range(0, len(clean_id), 2)]
        variants.update({':'.join(pairs), '-'.join(pairs)})
        variants.update({value.lower() for value in list(variants)})
    return sorted(value for value in variants if value)


def get_device_key_from_request() -> str:
    key = request.headers.get('X-Device-Key', '').strip().lower()
    if not re.fullmatch(r'[0-9a-f]{64}', key):
        return ''
    return key


def authenticate_device_key(device_id: str, allow_tofu: bool) -> bool:
    """Authenticate a device key, optionally registering it on first contact."""
    if device_status_collection is None:
        return False

    clean_id = normalize_device_id(device_id)
    raw_device_key = request.headers.get('X-Device-Key', '').strip()
    device_key = get_device_key_from_request()
    if not is_valid_device_id(clean_id):
        return False
    if raw_device_key and not device_key:
        return False

    existing = device_status_collection.find_one(
        {'deviceId': clean_id},
        {
            'deviceKeyHash': 1,
            'deviceKeyResetUntil': 1,
            'deviceKeyResetRequestedBy': 1,
        },
    )
    existing_hash = existing.get('deviceKeyHash') if existing else None
    if existing_hash:
        if not device_key:
            return False
        key_hash = hashlib.sha256(device_key.encode('ascii')).hexdigest()
        if secrets.compare_digest(str(existing_hash), key_hash):
            return True
        reset_until = existing.get('deviceKeyResetUntil')
        now = utcnow()
        if not allow_tofu or not reset_until or reset_until <= now:
            return False
        with get_device_write_lock(clean_id):
            fresh_now = utcnow()
            fresh_status = device_status_collection.find_one(
                {'deviceId': clean_id},
                {
                    'deviceKeyHash': 1,
                    'deviceKeyResetUntil': 1,
                    'deviceKeyResetRequestedBy': 1,
                },
            )
            reset_owner = fresh_status.get('deviceKeyResetRequestedBy') if fresh_status else None
            reset_until = fresh_status.get('deviceKeyResetUntil') if fresh_status else None
            fresh_hash = fresh_status.get('deviceKeyHash') if fresh_status else None
            if (
                not fresh_hash
                or not secrets.compare_digest(str(fresh_hash), str(existing_hash))
                or not reset_until
                or reset_until <= fresh_now
                or devices_collection is None
                or devices_collection.find_one({
                    'deviceId': clean_id,
                    'owner': reset_owner,
                    'claimed': True,
                }) is None
            ):
                return False
            rotated = device_status_collection.find_one_and_update(
                {
                    'deviceId': clean_id,
                    'deviceKeyHash': fresh_hash,
                    'deviceKeyResetUntil': {'$gt': fresh_now},
                    'deviceKeyResetRequestedBy': reset_owner,
                },
                {
                    '$set': {
                        'deviceKeyHash': key_hash,
                        'deviceKeyRegisteredAt': fresh_now,
                        'deviceKeyLastResetAt': fresh_now,
                        'deviceKeyLastResetBy': reset_owner,
                    },
                    '$unset': {
                        'deviceKeyResetUntil': '',
                        'deviceKeyResetRequestedAt': '',
                        'deviceKeyResetRequestedBy': '',
                    },
                },
                return_document=ReturnDocument.AFTER,
            )
        rotated_hash = rotated.get('deviceKeyHash') if rotated else None
        return bool(rotated_hash and secrets.compare_digest(str(rotated_hash), key_hash))
    if not device_key:
        return not DEVICE_AUTH_REQUIRED
    if not allow_tofu:
        return not DEVICE_AUTH_REQUIRED

    key_hash = hashlib.sha256(device_key.encode('ascii')).hexdigest()
    try:
        device_status_collection.update_one(
            {'deviceId': clean_id, 'deviceKeyHash': {'$exists': False}},
            {
                '$set': {
                    'deviceKeyHash': key_hash,
                    'deviceKeyRegisteredAt': utcnow(),
                },
                '$setOnInsert': {'deviceId': clean_id},
            },
            upsert=True,
        )
    except DuplicateKeyError:
        pass

    confirmed = device_status_collection.find_one({'deviceId': clean_id}, {'deviceKeyHash': 1})
    confirmed_hash = confirmed.get('deviceKeyHash') if confirmed else None
    return bool(confirmed_hash and secrets.compare_digest(str(confirmed_hash), key_hash))

def ensure_device_owner(device_id: str, user) -> bool:
    """检查设备是否属于当前用户"""
    if devices_collection is None or not user:
        return False
    owner = user.get('username')
    if not owner:
        return False
    clean_id = normalize_device_id(device_id)
    device = devices_collection.find_one({'deviceId': clean_id, 'owner': owner})
    return device is not None


def consume_valid_pairing_code(device_id: str, pairing_code: str):
    if pairing_codes_collection is None:
        return None
    clean_id = normalize_device_id(device_id)
    code = str(pairing_code or '').strip()
    now = utcnow()
    unlocked = {
        '$or': [
            {'lockedUntil': {'$exists': False}},
            {'lockedUntil': {'$lte': now}},
        ]
    }
    if re.fullmatch(r'\d{6}', code):
        consumed = pairing_codes_collection.find_one_and_delete({
            'deviceId': clean_id,
            'code': code,
            'expiresAt': {'$gt': now},
            **unlocked,
        })
        if consumed:
            return consumed

    failed = pairing_codes_collection.find_one_and_update(
        {
            'deviceId': clean_id,
            'expiresAt': {'$gt': now},
            **unlocked,
        },
        {'$inc': {'failedAttempts': 1}},
        return_document=ReturnDocument.AFTER,
    )
    if failed and failed.get('failedAttempts', 0) >= PAIRING_MAX_FAILED_ATTEMPTS:
        pairing_codes_collection.find_one_and_update(
            {
                'deviceId': clean_id,
                'failedAttempts': {'$gte': PAIRING_MAX_FAILED_ATTEMPTS},
            },
            {
                '$set': {
                    'failedAttempts': 0,
                    'lockedUntil': now + timedelta(seconds=PAIRING_LOCK_SECONDS),
                }
            },
            return_document=ReturnDocument.AFTER,
        )
    return None


def get_or_create_pairing_code(device_id: str):
    """Return one stable code even when status requests arrive concurrently."""
    if pairing_codes_collection is None:
        raise RuntimeError('Pairing code storage unavailable')

    clean_id = normalize_device_id(device_id)
    for _ in range(3):
        now = utcnow()
        existing = pairing_codes_collection.find_one({'deviceId': clean_id})
        if existing and existing.get('code') and existing.get('expiresAt') and existing['expiresAt'] > now:
            return existing

        replacement = {
            'code': f"{secrets.randbelow(900000) + 100000:06d}",
            'expiresAt': now + timedelta(hours=24),
            'createdAt': now,
            'failedAttempts': 0,
        }
        if existing is None:
            try:
                candidate = pairing_codes_collection.find_one_and_update(
                    {'deviceId': clean_id},
                    {'$setOnInsert': {'deviceId': clean_id, **replacement}},
                    upsert=True,
                    return_document=ReturnDocument.AFTER,
                )
            except DuplicateKeyError:
                candidate = None
        else:
            candidate = pairing_codes_collection.find_one_and_update(
                {
                    'deviceId': clean_id,
                    'code': existing.get('code'),
                    'expiresAt': existing.get('expiresAt'),
                },
                {'$set': replacement, '$unset': {'lockedUntil': ''}},
                return_document=ReturnDocument.AFTER,
            )

        if candidate and candidate.get('code') and candidate.get('expiresAt') and candidate['expiresAt'] > now:
            return candidate

    raise RuntimeError('Failed to allocate pairing code')


def claim_device_for_owner(device_id: str, owner: str, device_name: str, pairing_code: str):
    """Atomically consume a pairing code and claim only an unowned device."""
    clean_id = normalize_device_id(device_id)
    if not is_valid_device_id(clean_id):
        return None, 'Invalid deviceId format', 400
    with get_device_write_lock(clean_id):
        return _claim_device_for_owner_locked(clean_id, owner, device_name, pairing_code)


def _claim_device_for_owner_locked(clean_id: str, owner: str, device_name: str, pairing_code: str):
    if not owner:
        return None, 'Unauthorized', 401
    if devices_collection is None or pairing_codes_collection is None:
        return None, 'Database not connected', 500

    existing = devices_collection.find_one({'deviceId': clean_id})
    if existing and existing.get('claimed'):
        if existing.get('owner') != owner:
            return None, 'Device already claimed by another user', 403
        return None, 'Device is already claimed', 409

    if not consume_valid_pairing_code(clean_id, pairing_code):
        return None, 'Invalid or expired pairing code', 400

    # Always clear tenant-scoped remnants before assigning an unclaimed device.
    # deviceKeyHash is deliberately preserved by this helper.
    cleanup_device_artifacts(clean_id)

    now = utcnow()
    normalized_name = ((device_name or clean_id).strip() or clean_id)[:80]
    try:
        claimed = devices_collection.find_one_and_update(
            {
                'deviceId': clean_id,
                'claimed': {'$ne': True},
            },
            {
                '$set': {
                    'owner': owner,
                    'deviceName': normalized_name,
                    'claimed': True,
                    'updatedAt': now,
                },
                '$setOnInsert': {
                    'deviceId': clean_id,
                    'imageVersion': 0,
                    'activeContentMode': 'image',
                    'activeContentLabel': CONTENT_MODE_LABELS['image'],
                    'sleepIntervalSeconds': DEFAULT_SLEEP_INTERVAL_SECONDS,
                    'addedAt': now,
                    'createdAt': now,
                },
            },
            upsert=True,
            return_document=ReturnDocument.AFTER,
        )
    except DuplicateKeyError:
        claimed = None

    if claimed is None:
        current = devices_collection.find_one({'deviceId': clean_id})
        if current and current.get('claimed') and current.get('owner') != owner:
            return None, 'Device already claimed by another user', 403
        return None, 'Device claim conflict, request a new pairing code', 409
    # Remove a code that may have been regenerated by a simultaneous status call
    # after the original code was consumed but before the claim became visible.
    pairing_codes_collection.delete_one({'deviceId': clean_id})
    if device_status_collection is not None:
        device_status_collection.update_one(
            {'deviceId': clean_id},
            {
                '$set': {'everClaimed': True},
                '$unset': {'unclaimedExpiresAt': ''},
            },
        )
    return claimed, None, 200

# ==================== API: 用户注册 / 登录 ====================

@app.route('/api/auth/register', methods=['POST'])
def register():
    """用户注册"""
    global users_collection
    if users_collection is None:
        return jsonify({'success': False, 'error': 'Database not connected'}), 500

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()

    if not username or not password:
        return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400

    if len(username) < 3 or len(username) > 64 or len(password) < 8 or len(password) > 256:
        return jsonify({'success': False, 'error': '用户名或密码长度无效'}), 400

    first_account_only = not ALLOW_REGISTRATION
    if first_account_only and users_collection.find_one({}) is not None:
        return jsonify({'success': False, 'error': '注册已关闭，请联系管理员'}), 403
    if first_account_only:
        if len(ADMIN_BOOTSTRAP_TOKEN) < 32:
            return jsonify({'success': False, 'error': '管理员引导未配置'}), 503
        provided_bootstrap_token = str(
            request.headers.get('X-Admin-Bootstrap-Token')
            or data.get('bootstrapToken')
            or ''
        ).strip()
        if not secrets.compare_digest(provided_bootstrap_token, ADMIN_BOOTSTRAP_TOKEN):
            return jsonify({'success': False, 'error': '管理员引导凭据无效'}), 403

    try:
        user_doc = {
            'username': username,
            'passwordHash': hash_password(password),
            'createdAt': utcnow()
        }
        if first_account_only:
            # 稀疏唯一索引保证并发首次注册只有一个成功。
            user_doc['registrationSlot'] = 'first'
        users_collection.insert_one(user_doc)
        return jsonify({'success': True, 'message': '注册成功'})
    except DuplicateKeyError:
        return jsonify({'success': False, 'error': '用户名已存在或注册已关闭'}), 409
    except Exception as e:
        print(f'❌ Error registering user: {e}')
        return jsonify({'success': False, 'error': '注册失败'}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录，返回 token"""
    global users_collection
    if users_collection is None:
        return jsonify({'success': False, 'error': 'Database not connected'}), 500

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()

    if not username or not password:
        return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400
    if len(username) > 64 or len(password) > 256:
        return jsonify({'success': False, 'error': '用户名或密码长度无效'}), 400

    user = users_collection.find_one({'username': username})
    stored_password_hash = user.get('passwordHash', '') if user else DUMMY_PASSWORD_HASH
    valid_password, needs_migration = verify_password(stored_password_hash, password)
    if not stored_password_hash.startswith('scrypt:'):
        # Legacy SHA-256 and malformed hashes are fast to check; perform one
        # dummy scrypt verification so account existence is not exposed by a
        # large response-time difference. Valid legacy logins are migrated.
        check_password_hash(DUMMY_PASSWORD_HASH, password)
    if not user or not valid_password:
        return jsonify({'success': False, 'error': '用户名或密码错误'}), 400

    token = generate_token()
    token_expires_at = utcnow() + timedelta(seconds=AUTH_TOKEN_TTL_SECONDS)
    update_fields = {
        'tokenHash': hash_token(token),
        'tokenExpiresAt': token_expires_at,
        'lastLoginAt': utcnow(),
    }
    if needs_migration:
        update_fields['passwordHash'] = hash_password(password)
    users_collection.update_one(
        {'_id': user['_id']},
        {
            '$set': update_fields,
            '$unset': {'token': ''},
        }
    )

    return jsonify({
        'success': True,
        'token': token,
        'expiresAt': token_expires_at.isoformat() + 'Z',
        'user': {'username': username}
    })

@app.route('/api/auth/logout', methods=['POST'])
@login_required
def logout():
    """退出登录"""
    global users_collection
    user = getattr(request, 'user', None)
    if not user or users_collection is None:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    users_collection.update_one(
        {'_id': user['_id']},
        {'$unset': {'token': '', 'tokenHash': '', 'tokenExpiresAt': ''}}
    )
    return jsonify({'success': True, 'message': 'Logged out'})

@app.route('/api/auth/me', methods=['GET'])
@login_required
def me():
    """获取当前登录用户信息"""
    user = getattr(request, 'user', None)
    return jsonify({
        'success': True,
        'user': {
            'username': user.get('username')
        }
    })

# ==================== API: 设备管理 ====================

@app.route('/api/devices/preflight', methods=['POST'])
@login_required
def device_preflight():
    """添加设备前预检测：防止用户填错设备码。

    如果设备码从未查询过服务器，则拒绝添加。
    """
    try:
        data = request.get_json() or {}
        raw_id = (data.get('deviceId') or '').strip().upper()

        if not raw_id:
            return jsonify({'success': False, 'error': 'Missing deviceId'}), 400

        clean_id = raw_id.replace('-', '').replace(':', '')

        import re
        if not re.match(r'^[0-9A-F]{6}$|^[0-9A-F]{12}$', clean_id):
            return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400

        if devices_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        existing = devices_collection.find_one({'deviceId': clean_id})
        already_added = existing is not None

        has_seen_device = False
        last_seen_time = None
        recently_active = False

        if device_status_collection is not None:
            status = device_status_collection.find_one({'deviceId': clean_id})
            if status:
                has_seen_device = True
                last_seen_time = status.get('lastSeen', 0)
                recently_active = (time.time() * 1000 - last_seen_time) < 13 * 60 * 60 * 1000

        result = {
            'success': True,
            'deviceId': clean_id,
            'alreadyAdded': already_added,
            'deviceExists': has_seen_device,
            'recentlyActive': recently_active,
            'lastSeen': last_seen_time,
        }

        if already_added:
            result['message'] = f'设备 {clean_id} 已存在于您名下'
            result['canProceed'] = True
        elif has_seen_device:
            if recently_active:
                result['message'] = '设备码有效！该设备最近有活动记录'
                result['canProceed'] = True
            else:
                result['message'] = '该设备码历史上有活动记录，但最近未查询'
                result['canProceed'] = True
                result['warning'] = '设备码存在但最近不活跃，请确认是否正确'
        else:
            result['message'] = '未在服务器中找到该设备码的记录'
            result['canProceed'] = False
            result['error'] = f'设备码 {clean_id} 从未查询过本服务器，请检查设备码是否正确（或确认设备已先唤醒并配网）'

        return jsonify(result)
    except Exception as e:
        print(f'❌ Error in preflight check: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Device preflight failed'}), 500

@app.route('/api/devices/list', methods=['GET'])
@login_required
def get_devices_list():
    """获取当前用户的设备列表"""
    try:
        user = getattr(request, 'user', None)
        if devices_collection is None or not user:
            return jsonify({'success': True, 'devices': []})

        owner = user.get('username')
        devices = list(
            devices_collection.find({'owner': owner}, {'_id': 0})
            .sort('addedAt', -1)
        )
        return jsonify({'success': True, 'devices': devices})
    except Exception as e:
        print(f'❌ Error fetching devices: {e}')
        return jsonify({'success': False, 'error': 'Failed to fetch devices'}), 500

@app.route('/api/devices/add', methods=['POST'])
@login_required
def add_device():
    """使用设备端展示的一次性配对码添加设备。"""
    try:
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        data = request.get_json(silent=True) or {}
        device_id = str(data.get('deviceId') or '')
        if not device_id.strip():
            return jsonify({'success': False, 'error': 'Missing deviceId'}), 400
        clean_id = normalize_device_id(device_id)
        claimed, error, status = claim_device_for_owner(
            clean_id,
            owner,
            str(data.get('deviceName') or '').strip(),
            data.get('pairingCode'),
        )
        if claimed is None:
            return jsonify({'success': False, 'error': error}), status

        device = {key: value for key, value in claimed.items() if key != '_id'}
        for key in ('addedAt', 'createdAt', 'updatedAt'):
            if hasattr(device.get(key), 'isoformat'):
                device[key] = device[key].isoformat()
        print(f'✅ Device added: {clean_id}')
        return jsonify({'success': True, 'device': device})
    except Exception as e:
        print(f'❌ Error adding device: {e}')
        return jsonify({'success': False, 'error': 'Failed to add device'}), 500

@app.route('/api/devices/<device_id>', methods=['DELETE'])
@login_required
def delete_device(device_id):
    """删除当前用户的设备"""
    try:
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None

        clean_id = normalize_device_id(device_id)
        if not is_valid_device_id(clean_id):
            return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400
        if devices_collection is None or not owner:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        with get_device_write_lock(clean_id):
            result = devices_collection.delete_one({'deviceId': clean_id, 'owner': owner})
            if result.deleted_count == 0:
                return jsonify({'success': False, 'error': 'Device not found'}), 404
            cleanup_device_artifacts(clean_id)

        print(f'✅ Device deleted: {clean_id}')
        return jsonify({'success': True, 'message': 'Device deleted'})
    except Exception as e:
        print(f'❌ Error deleting device: {e}')
        return jsonify({'success': False, 'error': 'Failed to delete device'}), 500

@app.route('/api/devices', methods=['GET'])
@login_required
def get_devices_status():
    """获取当前用户的设备列表和状态"""
    try:
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None

        registered_devices = []
        if devices_collection is not None and owner:
            registered_devices = list(
                devices_collection.find({'owner': owner}, {'_id': 0})
            )

        devices = []
        for device in registered_devices:
            device_id = device['deviceId']
            content_meta = get_device_content_metadata(device)

            device_info = {
                'deviceId': device_id,
                'deviceName': device.get('deviceName', device_id),
                'addedAt': device.get('addedAt').isoformat() if hasattr(device.get('addedAt'), 'isoformat') else device.get('addedAt'),
                'online': False,  # Deep-sleep架构下设备通常离线
                'sleeping': False,  # Deep-sleep 架构：离线并不一定异常，后端给出“睡眠态”提示
                'claimed': device.get('claimed', False),
                'imageVersion': device.get('imageVersion', 0),
                'activeContentMode': content_meta['activeContentMode'],
                'activeTemplateId': content_meta['activeTemplateId'],
                'activeContentLabel': content_meta['activeContentLabel'],
                'sleepIntervalSeconds': content_meta['sleepIntervalSeconds'],
                'estimatedNextAutoWakeAt': None,
                'wakePolicyPending': False
            }

            # 检查设备最后活动时间
            if device_status_collection is not None:
                status = device_status_collection.find_one({'deviceId': device_id})
                if status:
                    last_seen = status.get('lastSeen', 0)
                    current_time = int(time.time() * 1000)
                    # Deep-sleep架构：最近5分钟内有活动则认为在线
                    device_info['online'] = (current_time - last_seen < 300000)
                    # Deep-sleep架构：在一个唤醒周期内无上报视为“睡眠中”，超过周期则视为“离线/失联”
                    # 默认周期：12小时唤醒一次；如果设备上报了云端动态间隔，则按该间隔再给 1 小时宽限
                    current_sleep_seconds = status.get('currentSleepSeconds')
                    if isinstance(current_sleep_seconds, (int, float)) and current_sleep_seconds > 0:
                        sleep_window_ms = int((current_sleep_seconds + 3600) * 1000)
                    else:
                        sleep_window_ms = 13 * 60 * 60 * 1000
                    device_info['sleeping'] = (not device_info['online']) and (current_time - last_seen < sleep_window_ms)
                    interval_for_estimate = content_meta['sleepIntervalSeconds']
                    content_updated_ms = to_epoch_ms(device.get('activeContentUpdatedAt') or device.get('updatedAt'))
                    if content_updated_ms and last_seen and content_updated_ms > last_seen:
                        if isinstance(current_sleep_seconds, (int, float)) and current_sleep_seconds > 0:
                            interval_for_estimate = int(current_sleep_seconds)
                            device_info['wakePolicyPending'] = True
                    if last_seen:
                        device_info['estimatedNextAutoWakeAt'] = int(last_seen + interval_for_estimate * 1000)
                    device_info['lastSeen'] = last_seen
                    device_info['lastWakeType'] = status.get('lastWakeType')
                    device_info['lastWakeCause'] = status.get('lastWakeCause')
                    device_info['lastManualWake'] = status.get('lastManualWake')
                    device_info['lastAutoWake'] = status.get('lastAutoWake')
                    device_info['ip'] = status.get('ip')
                    device_info['remoteIp'] = status.get('remoteIp')
                    device_info['rssi'] = status.get('rssi')
                    device_info['uptime_ms'] = status.get('uptime_ms')
                    device_info['freeHeap'] = status.get('freeHeap')
                    device_info['currentSleepSeconds'] = status.get('currentSleepSeconds')

            devices.append(device_info)

        return jsonify({'success': True, 'devices': devices})
    except Exception as e:
        print(f'❌ Error fetching device status: {e}')
        return jsonify({'success': False, 'error': 'Failed to fetch device status'}), 500

# ==================== API: 设备绑定状态查询和绑定 ====================

@app.route('/api/device/status', methods=['POST'])
def device_status():
    """设备查询绑定状态（无需登录，设备调用）

    返回：
    - claimed: 是否已绑定
    - imageVersion: 最新图片版本号
    - imageUrl: 图片下载URL（仅已绑定且有图片时返回）
    - pairingCode: 配对码（仅未绑定时返回）
    """
    try:
        if request.content_length is not None and request.content_length > DEVICE_STATUS_MAX_BODY_BYTES:
            return jsonify({'success': False, 'error': 'Device status payload too large'}), 413
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({'success': False, 'error': 'Invalid JSON payload'}), 400
        device_id = str(data.get('deviceId') or '')

        if not device_id:
            return jsonify({'success': False, 'error': 'Missing deviceId'}), 400

        clean_id = normalize_device_id(device_id)
        if not is_valid_device_id(clean_id):
            return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400
        if devices_collection is None or device_status_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500
        if not authenticate_device_key(clean_id, allow_tofu=True):
            return jsonify({'success': False, 'error': 'Invalid device credentials'}), 401

        telemetry = {
            'lastSeen': int(time.time() * 1000),
            'updatedAt': utcnow()
        }
        now_ms = telemetry['lastSeen']

        forwarded_for = request.headers.get('X-Forwarded-For', '')
        if forwarded_for:
            telemetry['remoteIp'] = sanitize_telemetry_text(
                forwarded_for.split(',')[0], 64,
            )
        elif request.remote_addr:
            telemetry['remoteIp'] = sanitize_telemetry_text(request.remote_addr, 64)

        ip = data.get('ip')
        if isinstance(ip, str) and ip.strip():
            telemetry['ip'] = sanitize_telemetry_text(ip, 64)

        for field, (minimum, maximum) in TELEMETRY_NUMERIC_LIMITS.items():
            value = data.get(field)
            if (
                isinstance(value, (int, float))
                and not isinstance(value, bool)
                and (isinstance(value, int) or math.isfinite(value))
            ):
                telemetry[field] = min(max(int(value), minimum), maximum)

        raw_wake_type = data.get('wakeType')
        raw_wake_cause = data.get('wakeCause')
        wake_type = raw_wake_type.strip().lower()[:16] if isinstance(raw_wake_type, str) else ''
        wake_cause = sanitize_telemetry_text(raw_wake_cause, 64)
        if wake_type in ('manual', 'auto', 'reset', 'other'):
            telemetry['lastWakeType'] = wake_type
            if wake_type == 'manual':
                telemetry['lastManualWake'] = now_ms
            elif wake_type == 'auto':
                telemetry['lastAutoWake'] = now_ms
        if wake_cause:
            telemetry['lastWakeCause'] = wake_cause

        print(
            f"📡 设备 {clean_id} 上报: "
            f"ip={telemetry.get('ip', '-')}, "
            f"remoteIp={telemetry.get('remoteIp', '-')}, "
            f"rssi={telemetry.get('rssi', '-')}, "
            f"uptime_ms={telemetry.get('uptime_ms', '-')}, "
            f"freeHeap={telemetry.get('freeHeap', '-')}, "
            f"wakeType={telemetry.get('lastWakeType', '-')}, "
            f"wakeCause={telemetry.get('lastWakeCause', '-')}, "
            f"currentSleepSeconds={telemetry.get('currentSleepSeconds', '-')}"
        )

        device = devices_collection.find_one({'deviceId': clean_id})
        claimed = device is not None and device.get('claimed', False)

        # 只给“从未绑定过”的 TOFU 记录设置滑动 TTL，限制任意设备码
        # 导致的永久数据库增长。曾绑定设备必须永久保留密钥哈希，避免
        # 解绑后重新抢占。
        status_update = {'$set': telemetry}
        if claimed:
            telemetry['everClaimed'] = True
            status_update['$unset'] = {'unclaimedExpiresAt': ''}
        else:
            current_status = device_status_collection.find_one(
                {'deviceId': clean_id}, {'everClaimed': 1},
            )
            if current_status and current_status.get('everClaimed'):
                status_update['$unset'] = {'unclaimedExpiresAt': ''}
            else:
                telemetry['unclaimedExpiresAt'] = (
                    utcnow() + timedelta(seconds=UNCLAIMED_DEVICE_TTL_SECONDS)
                )
        device_status_collection.update_one(
            {'deviceId': clean_id}, status_update, upsert=True,
        )

        response = {
            'success': True,
            'deviceId': clean_id,
            'claimed': claimed
        }

        if claimed and device:
            # ===== 模板设备按需渲染 =====
            # 设备唤醒时，如果内容过期则同步渲染新数据，确保设备拿到最新内容
            template_id = device.get('activeTemplateId')
            template_config = device.get('templateConfig', {})
            content_mode = device.get('activeContentMode', 'image')

            if content_mode == 'template' and template_id and isinstance(template_config, dict):
                try:
                    with get_device_write_lock(clean_id):
                        latest_device = devices_collection.find_one({'deviceId': clean_id, 'claimed': True})
                        if not latest_device:
                            raise RuntimeError('Device ownership changed during rendering')
                        latest_template_id = latest_device.get('activeTemplateId')
                        latest_template_config = latest_device.get('templateConfig', {})
                        latest_mode = latest_device.get('activeContentMode', 'image')
                        should_render = (
                            latest_mode == 'template'
                            and latest_template_id == template_id
                            and isinstance(latest_template_config, dict)
                            and _should_re_render_template(
                                latest_template_id, latest_template_config, latest_device, wake_type
                            )
                        )
                        if not should_render:
                            device = latest_device
                        else:
                            print(f'🔄 设备唤醒触发渲染: {clean_id}, 模板={template_id}')
                            epd_data = render_template_epd_data(template_id, latest_template_config)
                            if not save_device_image(clean_id, epd_data):
                                raise RuntimeError('Failed to save rendered image')
                            device = devices_collection.find_one_and_update(
                                {
                                    'deviceId': clean_id,
                                    'claimed': True,
                                    'activeContentMode': 'template',
                                    'activeTemplateId': template_id,
                                },
                                {
                                    '$inc': {'imageVersion': 1},
                                    '$set': {
                                        'imageSizeChars': len(epd_data),
                                        'imageSizeBytes': len(epd_data.encode('utf-8')),
                                        'imageSha256': hashlib.sha256(epd_data.encode('utf-8')).hexdigest(),
                                        'renderSource': 'pillow',
                                        'activeContentUpdatedAt': utcnow(),
                                        'updatedAt': utcnow(),
                                    },
                                },
                                return_document=ReturnDocument.AFTER,
                            )
                            if device is None:
                                raise RuntimeError('Device content changed during rendering')
                            print(f"✅ 唤醒渲染完成: {clean_id}, 新版本={device.get('imageVersion')}")
                except Exception as e:
                    print(f'⚠️ 唤醒渲染失败: {clean_id} -> {e}（设备将使用旧数据）')

            # 已绑定：返回图片版本和下载URL
            image_version = device.get('imageVersion', 0)
            content_meta = get_device_content_metadata(device)
            response['imageVersion'] = image_version
            response['nextSleepSeconds'] = content_meta['sleepIntervalSeconds']

            # 检查是否有持久化的图片
            image_path = get_ready_device_image_path(clean_id, device)
            if image_path is not None and image_version > 0:
                # 构建稳定的下载URL
                response['imageUrl'] = build_raw_image_url(clean_id, image_version)
                # 返回云端侧元数据，设备可做轻量校验（不强制）
                if device.get('imageSizeChars') is not None:
                    response['imageSizeChars'] = device.get('imageSizeChars')
                if device.get('imageSha256') is not None:
                    response['imageSha256'] = device.get('imageSha256')

            print(f"📊 设备 {clean_id} 查询状态: claimed=True, imageVersion={image_version}, "
                  f"nextSleepSeconds={content_meta['sleepIntervalSeconds']}")
        else:
            # 未绑定：生成或返回配对码
            response['imageVersion'] = 0
            response['nextSleepSeconds'] = DEFAULT_SLEEP_INTERVAL_SECONDS

            pairing_doc = get_or_create_pairing_code(clean_id)
            pairing_code = pairing_doc['code']
            expires_at = pairing_doc['expiresAt']

            if expires_at:
                expires_in = int((expires_at - utcnow()).total_seconds())
                if expires_in < 0:
                    expires_in = 0
            else:
                expires_in = 86400

            response['pairingCode'] = pairing_code
            response['expiresIn'] = expires_in

            print(f'📊 设备 {clean_id} 查询状态: claimed=False, pairingCode=已生成')

        return jsonify(response)
    except Exception as e:
        print(f'❌ Error querying device status: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Failed to query device status'}), 500

@app.route('/api/device/claim', methods=['POST'])
@login_required
def device_claim():
    """用户绑定设备（需要登录）"""
    try:
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None

        if not owner:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401

        data = request.get_json(silent=True) or {}
        device_id = str(data.get('deviceId') or '')

        if not device_id:
            return jsonify({'success': False, 'error': 'Missing deviceId'}), 400

        clean_id = normalize_device_id(device_id)
        claimed, error, status = claim_device_for_owner(
            clean_id,
            owner,
            str(data.get('deviceName') or '').strip(),
            data.get('pairingCode'),
        )
        if claimed is None:
            return jsonify({'success': False, 'error': error}), status

        print(f'✅ Device claimed: {clean_id} by {owner}')

        return jsonify({
            'success': True,
            'message': 'Device claimed successfully',
            'deviceId': clean_id
        })
    except Exception as e:
        print(f'❌ Error claiming device: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Failed to claim device'}), 500

@app.route('/api/device/unbind', methods=['POST'])
@login_required
def device_unbind():
    """解绑设备（需要登录，仅限设备所有者）"""
    try:
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None

        if not owner:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401

        data = request.get_json(silent=True) or {}
        device_id = str(data.get('deviceId') or '')

        if not device_id:
            return jsonify({'success': False, 'error': 'Missing deviceId'}), 400

        clean_id = normalize_device_id(device_id)
        if not is_valid_device_id(clean_id):
            return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400

        if devices_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        with get_device_write_lock(clean_id):
            result = devices_collection.update_one(
                {'deviceId': clean_id, 'owner': owner},
                {
                    '$set': {
                        'claimed': False,
                        'imageVersion': 0,
                        'activeContentMode': 'image',
                        'activeContentLabel': CONTENT_MODE_LABELS['image'],
                        'sleepIntervalSeconds': DEFAULT_SLEEP_INTERVAL_SECONDS,
                        'updatedAt': utcnow(),
                    },
                    '$unset': {
                        'owner': '', 'activeTemplateId': '', 'templateConfig': '',
                        'imageSizeChars': '', 'imageSizeBytes': '', 'imageSha256': '',
                        'renderSource': '', 'activeContentUpdatedAt': '',
                        'nameplateName': '', 'nameplateBatchId': '',
                    },
                },
            )
            if result.matched_count == 0:
                return jsonify({'success': False, 'error': 'Device not found or no permission'}), 404
            cleanup_device_artifacts(clean_id)

        print(f'✅ Device unbound: {clean_id} by {owner}')

        return jsonify({
            'success': True,
            'message': 'Device unbound successfully',
            'deviceId': clean_id
        })
    except Exception as e:
        print(f'❌ Error unbinding device: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Failed to unbind device'}), 500


@app.route('/api/device/auth/reset', methods=['POST'])
@login_required
def reset_device_credentials():
    """Open one short, owner-authorized window for replacing a lost device key."""
    data = request.get_json(silent=True) or {}
    clean_id = normalize_device_id(str(data.get('deviceId') or ''))
    user = getattr(request, 'user', None)
    owner = user.get('username') if user else None
    if not is_valid_device_id(clean_id):
        return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400
    if devices_collection is None or device_status_collection is None:
        return jsonify({'success': False, 'error': 'Database not connected'}), 500

    now = utcnow()
    reset_until = now + timedelta(seconds=DEVICE_KEY_RESET_WINDOW_SECONDS)
    with get_device_write_lock(clean_id):
        current_device = devices_collection.find_one({
            'deviceId': clean_id,
            'owner': owner,
            'claimed': True,
        })
        if current_device is None:
            return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403
        device_status_collection.update_one(
            {'deviceId': clean_id},
            {
                '$set': {
                    'deviceKeyResetUntil': reset_until,
                    'deviceKeyResetRequestedAt': now,
                    'deviceKeyResetRequestedBy': owner,
                },
                '$setOnInsert': {'deviceId': clean_id},
            },
            upsert=True,
        )
    return jsonify({
        'success': True,
        'deviceId': clean_id,
        'resetUntil': reset_until.isoformat() + 'Z',
    })

# ==================== API: 页面管理 ====================

@app.route('/api/pages/list/<device_id>', methods=['GET'])
@login_required
def get_pages(device_id):
    """获取设备的所有页面"""
    try:
        user = getattr(request, 'user', None)
        clean_id = normalize_device_id(device_id)
        if not ensure_device_owner(clean_id, user):
            return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

        if pages_collection is None:
            return jsonify({'success': True, 'pages': []})

        # 兼容历史数据：deviceId 可能未规范化写入（带分隔符/小写等）
        candidates = device_id_variants(device_id)

        # 列表接口仅返回轻量字段，避免把 data.imageData（base64）整包带回导致前端卡顿/不显示
        limit = request.args.get('limit', '200')
        try:
            limit = int(limit)
        except Exception:
            limit = 200
        limit = max(1, min(limit, 500))

        pages = list(pages_collection.find(
            {
                'deviceId': {'$in': candidates},
                '$or': [
                    {'owner': user.get('username')},
                    {'owner': {'$exists': False}},
                    {'owner': None},
                ],
            },
            {
                '_id': 0,
                'pageId': 1,
                'deviceId': 1,
                'name': 1,
                'type': 1,
                'thumbnail': 1,
                'createdAt': 1,
                'updatedAt': 1
            }
        ).sort('updatedAt', -1).limit(limit))

        for page in pages:
            if hasattr(page.get('createdAt'), 'isoformat'):
                page['createdAt'] = page['createdAt'].isoformat()
            if hasattr(page.get('updatedAt'), 'isoformat'):
                page['updatedAt'] = page['updatedAt'].isoformat()

        return jsonify({'success': True, 'pages': pages})
    except Exception as e:
        print(f'❌ Error fetching pages: {e}')
        return jsonify({'success': False, 'error': 'Failed to fetch pages'}), 500

@app.route('/api/pages/save', methods=['POST'])
@login_required
def save_page():
    """保存页面"""
    try:
        data = request.get_json(silent=True) or {}
        device_id = data.get('deviceId')
        page_id = data.get('pageId')
        page_name = data.get('name', '未命名页面')
        page_type = data.get('type', 'custom')
        page_data = data.get('data', {})
        thumbnail = data.get('thumbnail', '')

        if not device_id:
            return jsonify({'success': False, 'error': 'Missing deviceId'}), 400

        clean_id = normalize_device_id(device_id)
        if not is_valid_device_id(clean_id):
            return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400
        if page_id is not None and not re.fullmatch(r'[A-Za-z0-9_-]{1,64}', str(page_id)):
            return jsonify({'success': False, 'error': 'Invalid pageId format'}), 400
        if not isinstance(page_name, str) or not 1 <= len(page_name.strip()) <= PAGE_MAX_NAME_CHARS:
            return jsonify({'success': False, 'error': 'Invalid page name'}), 400
        if not isinstance(page_type, str) or not 1 <= len(page_type) <= PAGE_MAX_TYPE_CHARS:
            return jsonify({'success': False, 'error': 'Invalid page type'}), 400
        if not isinstance(page_data, dict):
            return jsonify({'success': False, 'error': 'Page data must be an object'}), 400
        if not isinstance(thumbnail, str) or len(thumbnail.encode('utf-8')) > PAGE_MAX_THUMBNAIL_BYTES:
            return jsonify({'success': False, 'error': 'Page thumbnail is too large'}), 413
        if len(json.dumps(page_data, ensure_ascii=False).encode('utf-8')) > PAGE_MAX_DATA_BYTES:
            return jsonify({'success': False, 'error': 'Page data is too large'}), 413
        page_name = page_name.strip()
        page_id = str(page_id) if page_id is not None else None

        user = getattr(request, 'user', None)
        if not ensure_device_owner(clean_id, user):
            return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

        if pages_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        now = utcnow()
        owner = user.get('username') if user else None

        if page_id:
            legacy_device_ids = device_id_variants(device_id)
            result = pages_collection.update_one(
                {
                    'pageId': page_id,
                    'deviceId': {'$in': legacy_device_ids},
                    '$or': [
                        {'owner': owner},
                        {'owner': {'$exists': False}},
                    ],
                },
                {'$set': {
                    'deviceId': clean_id,
                    'name': page_name,
                    'type': page_type,
                    'data': page_data,
                    'thumbnail': thumbnail,
                    'owner': owner,
                    'updatedAt': now
                }}
            )
            if result.matched_count == 0:
                return jsonify({'success': False, 'error': 'Page not found'}), 404

            print(f'✅ Page updated: {page_id}')
        else:
            import uuid
            page_id = uuid.uuid4().hex

            page = {
                'pageId': page_id,
                'deviceId': clean_id,
                'name': page_name,
                'type': page_type,
                'data': page_data,
                'thumbnail': thumbnail,
                'owner': owner,
                'createdAt': now,
                'updatedAt': now
            }
            pages_collection.insert_one(page)
            print(f'✅ Page created: {page_id}')

        return jsonify({
            'success': True,
            'pageId': page_id,
            'message': 'Page saved'
        })
    except Exception as e:
        print(f'❌ Error saving page: {e}')
        return jsonify({'success': False, 'error': 'Failed to save page'}), 500

@app.route('/api/pages/<page_id>', methods=['GET'])
@login_required
def get_page(page_id):
    """获取单个页面详情"""
    try:
        if pages_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        if not re.fullmatch(r'[A-Za-z0-9_-]{1,64}', str(page_id)):
            return jsonify({'success': False, 'error': 'Page not found'}), 404
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        page = find_page_for_owner(page_id, owner, {'_id': 0})
        if not page:
            return jsonify({'success': False, 'error': 'Page not found'}), 404

        device_id = page.get('deviceId')
        if (
            not is_valid_device_id(normalize_device_id(device_id))
            or (page.get('owner') is not None and page.get('owner') != owner)
            or not ensure_device_owner(device_id, user)
        ):
            return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

        if hasattr(page.get('createdAt'), 'isoformat'):
            page['createdAt'] = page['createdAt'].isoformat()
        if hasattr(page.get('updatedAt'), 'isoformat'):
            page['updatedAt'] = page['updatedAt'].isoformat()

        return jsonify({'success': True, 'page': page})
    except Exception as e:
        print(f'❌ Error fetching page: {e}')
        return jsonify({'success': False, 'error': 'Failed to fetch page'}), 500

@app.route('/api/pages/<page_id>', methods=['DELETE'])
@login_required
def delete_page(page_id):
    """删除页面"""
    try:
        if pages_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        if not re.fullmatch(r'[A-Za-z0-9_-]{1,64}', str(page_id)):
            return jsonify({'success': False, 'error': 'Page not found'}), 404
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        page = find_page_for_owner(page_id, owner)
        if not page:
            return jsonify({'success': False, 'error': 'Page not found'}), 404

        device_id = page.get('deviceId')
        if (
            not is_valid_device_id(normalize_device_id(device_id))
            or (page.get('owner') is not None and page.get('owner') != owner)
            or not ensure_device_owner(device_id, user)
        ):
            return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

        delete_filter = {'_id': page['_id']} if page.get('_id') is not None else {
            'pageId': page_id,
            'deviceId': device_id,
            'owner': page.get('owner'),
        }
        result = pages_collection.delete_one(delete_filter)
        if result.deleted_count == 0:
            return jsonify({'success': False, 'error': 'Page not found'}), 404

        print(f'✅ Page deleted: {page_id}')
        return jsonify({'success': True, 'message': 'Page deleted'})
    except Exception as e:
        print(f'❌ Error deleting page: {e}')
        return jsonify({'success': False, 'error': 'Failed to delete page'}), 500

# ==================== API: 模板 ====================

TEMPLATES = [
    {
        'templateId': 'weather',
        'name': '天气',
        'icon': '🌤️',
        'description': '显示天气信息',
        'preview': '/templates/weather.png',
        'defaultData': {
            'type': 'template',
            'template': 'weather',
            'city': '',
            'showForecast': True
        }
    },
    {
        'templateId': 'calendar',
        'name': '日历',
        'icon': '📅',
        'description': '显示日历和日程',
        'preview': '/templates/calendar.png',
        'defaultData': {
            'type': 'template',
            'template': 'calendar',
            'showEvents': True
        }
    },
    {
        'templateId': 'todo',
        'name': '待办事项',
        'icon': '✅',
        'description': '显示待办事项列表',
        'preview': '/templates/todo.png',
        'defaultData': {
            'type': 'template',
            'template': 'todo',
            'items': []
        }
    },
    {
        'templateId': 'quote',
        'name': '每日一言',
        'icon': '💬',
        'description': '显示励志名言或诗词',
        'preview': '/templates/quote.png',
        'defaultData': {
            'type': 'template',
            'template': 'quote',
            'category': 'motivational'
        }
    },
    {
        'templateId': 'qrcode',
        'name': '二维码',
        'icon': '📱',
        'description': '显示自定义二维码',
        'preview': '/templates/qrcode.png',
        'defaultData': {
            'type': 'template',
            'template': 'qrcode',
            'content': '',
            'title': ''
        }
    },
    {
        'templateId': 'nameplate',
        'name': '会议名牌',
        'icon': 'meeting-nameplate',
        'description': 'Pheno 品牌姓名牌',
        'preview': '/templates/nameplate.png',
        'defaultData': {
            'type': 'template',
            'template': 'nameplate',
            'name': '',
            'backgroundStyle': 'formal_red',
            'title': '',
            'subtitle': ''
        }
    }
]
ACTIVE_TEMPLATE_IDS = {'nameplate'}

def get_active_templates():
    return [template for template in TEMPLATES if template.get('templateId') in ACTIVE_TEMPLATE_IDS]

@app.route('/api/templates', methods=['GET'])
def get_templates():
    """获取所有可用模板"""
    return jsonify({'success': True, 'templates': get_active_templates()})

@app.route('/api/templates/<template_id>', methods=['GET'])
def get_template(template_id):
    """获取单个模板详情"""
    if template_id not in ACTIVE_TEMPLATE_IDS:
        return jsonify({'success': False, 'error': 'Template not available'}), 404
    template = next((t for t in TEMPLATES if t['templateId'] == template_id), None)
    if not template:
        return jsonify({'success': False, 'error': 'Template not found'}), 404
    return jsonify({'success': True, 'template': template})


@app.route('/api/nameplate/templates', methods=['GET'])
@login_required
def list_saved_nameplate_templates():
    """列出当前账号保存的会议名牌模板。"""
    try:
        if saved_nameplate_templates_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        if not owner:
            return jsonify({'success': False, 'error': 'Missing owner'}), 401

        docs = list(saved_nameplate_templates_collection.find(
            {'owner': owner, 'baseTemplateId': 'nameplate'},
            {'_id': 0}
        ).sort('updatedAt', -1))

        return jsonify({
            'success': True,
            'templates': [serialize_saved_nameplate_template(doc) for doc in docs]
        })
    except Exception as e:
        print(f'❌ Error listing saved nameplate templates: {e}')
        return jsonify({'success': False, 'error': 'Failed to list templates'}), 500


@app.route('/api/nameplate/templates/<template_id>', methods=['GET'])
@login_required
def get_saved_nameplate_template(template_id):
    """获取当前账号保存的单个会议名牌模板。"""
    try:
        if saved_nameplate_templates_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        doc = saved_nameplate_templates_collection.find_one(
            {'owner': owner, 'templateId': str(template_id), 'baseTemplateId': 'nameplate'},
            {'_id': 0}
        )
        if not doc:
            return jsonify({'success': False, 'error': 'Template not found'}), 404

        return jsonify({'success': True, 'template': serialize_saved_nameplate_template(doc)})
    except Exception as e:
        print(f'❌ Error fetching saved nameplate template: {e}')
        return jsonify({'success': False, 'error': 'Failed to fetch template'}), 500


@app.route('/api/nameplate/templates', methods=['POST'])
@login_required
def save_saved_nameplate_template():
    """新建或更新当前账号保存的会议名牌模板。"""
    try:
        if saved_nameplate_templates_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        data = request.get_json() or {}
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        if not owner:
            return jsonify({'success': False, 'error': 'Missing owner'}), 401

        template_id = str(data.get('templateId') or '').strip()
        name = normalize_nameplate_template_name(data.get('name'))
        template_config = normalize_nameplate_template_config(data.get('templateConfig'))
        now = utcnow()

        if template_id:
            result = saved_nameplate_templates_collection.update_one(
                {'owner': owner, 'templateId': template_id, 'baseTemplateId': 'nameplate'},
                {'$set': {
                    'name': name,
                    'templateConfig': template_config,
                    'updatedAt': now,
                }}
            )
            if result.matched_count == 0:
                return jsonify({'success': False, 'error': 'Template not found'}), 404
        else:
            template_id = secrets.token_hex(6)
            doc = {
                'templateId': template_id,
                'owner': owner,
                'baseTemplateId': 'nameplate',
                'name': name,
                'templateConfig': template_config,
                'createdAt': now,
                'updatedAt': now,
            }
            saved_nameplate_templates_collection.insert_one(doc)

        saved = saved_nameplate_templates_collection.find_one(
            {'owner': owner, 'templateId': template_id, 'baseTemplateId': 'nameplate'},
            {'_id': 0}
        )
        return jsonify({'success': True, 'template': serialize_saved_nameplate_template(saved)})
    except DuplicateKeyError:
        return jsonify({'success': False, 'error': 'Template id conflict'}), 409
    except Exception as e:
        print(f'❌ Error saving nameplate template: {e}')
        return jsonify({'success': False, 'error': 'Failed to save template'}), 500


@app.route('/api/nameplate/templates/<template_id>', methods=['DELETE'])
@login_required
def delete_saved_nameplate_template(template_id):
    """删除当前账号保存的会议名牌模板。"""
    try:
        if saved_nameplate_templates_collection is None:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None
        result = saved_nameplate_templates_collection.delete_one(
            {'owner': owner, 'templateId': str(template_id), 'baseTemplateId': 'nameplate'}
        )
        if result.deleted_count == 0:
            return jsonify({'success': False, 'error': 'Template not found'}), 404
        return jsonify({'success': True})
    except Exception as e:
        print(f'❌ Error deleting saved nameplate template: {e}')
        return jsonify({'success': False, 'error': 'Failed to delete template'}), 500

# ==================== API: EPD 控制（HTTP拉取架构） ====================

@app.route('/api/epd/init', methods=['POST'])
@login_required
def epd_init():
    """初始化 EPD（Deep-sleep架构下此接口仅用于记录，不直接控制设备）"""
    data = request.get_json(silent=True) or {}
    device_id = data.get('deviceId')
    epd_type = data.get('epdType')

    if not device_id or epd_type is None:
        return jsonify({'success': False, 'error': 'Missing deviceId or epdType'}), 400

    user = getattr(request, 'user', None)
    if not ensure_device_owner(device_id, user):
        return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

    clean_id = normalize_device_id(device_id)
    print(f'📱 EPD init recorded for {clean_id}, type={epd_type}')
    return jsonify({'success': True, 'message': 'EPD init recorded (device will apply on next wake)'})

@app.route('/api/epd/load', methods=['POST'])
@login_required
def epd_load():
    """上传图片数据（持久化保存，设备下次唤醒时拉取）"""
    data = request.get_json(silent=True) or {}
    device_id = data.get('deviceId')
    image_data = data.get('data')
    content_meta = build_content_metadata(data.get('contentMode'), data.get('templateId'))

    if not device_id or not image_data:
        return jsonify({'success': False, 'error': 'Missing deviceId or data'}), 400

    user = getattr(request, 'user', None)
    if not ensure_device_owner(device_id, user):
        return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

    clean_id = normalize_device_id(device_id)

    # 发布阶段就做完整性校验：长度/字符集不符合直接拒绝
    ok, err = validate_epd_text_payload(image_data)
    if not ok:
        print(f'❌ 发布数据校验失败: {clean_id} -> {err}')
        return jsonify({'success': False, 'error': f'Invalid EPD data: {err}'}), 400

    # 计算元数据（用于 status 提示 / 设备轻量校验）
    image_size_chars = len(image_data)
    image_size_bytes = len(image_data.encode('utf-8'))
    image_sha256 = hashlib.sha256(image_data.encode('utf-8')).hexdigest()

    if devices_collection is not None:
        now = utcnow()
        update_fields = {
            'imageSizeChars': image_size_chars,
            'imageSizeBytes': image_size_bytes,
            'imageSha256': image_sha256,
            'activeContentMode': content_meta['activeContentMode'],
            'activeContentLabel': content_meta['activeContentLabel'],
            'sleepIntervalSeconds': content_meta['sleepIntervalSeconds'],
            'activeContentUpdatedAt': now,
            'updatedAt': now
        }
        update_doc = {'$inc': {'imageVersion': 1}, '$set': update_fields}
        if content_meta['activeTemplateId']:
            update_fields['activeTemplateId'] = content_meta['activeTemplateId']
        else:
            update_doc['$unset'] = {'activeTemplateId': ''}

        owner = user.get('username') if user else None
        with get_device_write_lock(clean_id):
            current_device = devices_collection.find_one({
                'deviceId': clean_id, 'owner': owner, 'claimed': True,
            })
            if current_device is None:
                return jsonify({'success': False, 'error': 'Device changed during publish'}), 409
            if not save_device_image(clean_id, image_data):
                return jsonify({'success': False, 'error': 'Failed to save image'}), 500
            updated_device = devices_collection.find_one_and_update(
                {'deviceId': clean_id, 'owner': owner, 'claimed': True},
                update_doc,
                return_document=ReturnDocument.AFTER,
            )
        if updated_device is None:
            return jsonify({'success': False, 'error': 'Device changed during publish'}), 409
        new_version = updated_device.get('imageVersion', 0)

        print(f'✅ 图片已保存: {clean_id}, 新版本: {new_version}')
        print(f'   数据大小: {len(image_data)} 字符 ({len(image_data)/1024:.2f} KB)')
        print(f"   当前内容: {content_meta['activeContentLabel']}, "
              f"唤醒间隔: {content_meta['sleepIntervalSeconds']} 秒")
        print(f'   设备下次唤醒时将自动拉取更新')

        return jsonify({
            'success': True,
            'message': 'Image saved, device will update on next wake',
            'imageVersion': new_version,
            'imageUrl': build_raw_image_url(clean_id, new_version),
            'imageSizeChars': image_size_chars,
            'imageSha256': image_sha256,
            'activeContentLabel': content_meta['activeContentLabel'],
            'sleepIntervalSeconds': content_meta['sleepIntervalSeconds']
        })

    return jsonify({'success': True, 'message': 'Image saved'})

@app.route('/api/epd/raw/<device_id>', methods=['GET'])
def epd_raw_download(device_id):
    """下载设备的原始图片数据（ESP32通过HTTP下载）

    返回 text/plain 格式的 a~p 编码字符串
    """
    clean_id = normalize_device_id(device_id)
    if not is_valid_device_id(clean_id):
        return jsonify({'success': False, 'error': 'Invalid deviceId format'}), 400
    if not authenticate_device_key(clean_id, allow_tofu=False):
        return jsonify({'success': False, 'error': 'Invalid device credentials'}), 401

    image_path = get_device_image_path(clean_id, create_parent=False)
    if not image_path.is_file():
        print(f'❌ 图片不存在: {clean_id}')
        return jsonify({'error': 'Image not found'}), 404

    device = None
    if devices_collection is not None:
        device = devices_collection.find_one(
            {'deviceId': clean_id},
            {'_id': 0, 'imageSha256': 1, 'imageSizeChars': 1},
        )
    if get_ready_device_image_path(clean_id, device) is None:
        print(f'❌ 存储图片或元数据校验失败: {clean_id}')
        return jsonify({'success': False, 'error': 'Stored image is invalid'}), 503

    data_size_bytes = image_path.stat().st_size
    print(f'📥 ESP32下载图片: {clean_id}')
    print(f'   文件大小: {data_size_bytes} 字节 ({data_size_bytes/1024:.2f} KB)')

    if data_size_bytes != EPD_EXPECTED_CHARS:
        print(f'❌ 数据大小不匹配: 期望 {EPD_EXPECTED_CHARS}, 实际 {data_size_bytes}')
        return jsonify({'success': False, 'error': 'Stored image is invalid'}), 503

    # 用 send_file 直接流式发送文件，减少内存占用，并提供条件请求/ETag（更利于代理/断点续传扩展）
    resp = send_file(
        image_path,
        mimetype='text/plain',
        conditional=True,
        etag=True,
        max_age=0
    )
    resp.headers['Content-Type'] = 'text/plain; charset=utf-8'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    resp.headers['Accept-Ranges'] = 'bytes'
    resp.headers['X-EPD-Expected-Chars'] = str(EPD_EXPECTED_CHARS)

    # 如果 DB 里有 hash，就带上；没有也不强制（兼容旧数据）
    if device:
        if device.get('imageSha256'):
            resp.headers['X-EPD-SHA256'] = device['imageSha256']
        if device.get('imageSizeChars') is not None:
            resp.headers['X-EPD-Chars'] = str(device['imageSizeChars'])

    return resp

@app.route('/api/epd/show', methods=['POST'])
@login_required
def epd_show():
    """触发设备显示（Deep-sleep架构下此接口仅用于记录）"""
    data = request.get_json(silent=True) or {}
    device_id = data.get('deviceId')

    if not device_id:
        return jsonify({'success': False, 'error': 'Missing deviceId'}), 400

    user = getattr(request, 'user', None)
    if not ensure_device_owner(device_id, user):
        return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

    clean_id = normalize_device_id(device_id)
    print(f'📺 Show command recorded for {clean_id} (device will display on next wake)')
    return jsonify({'success': True, 'message': 'Show command recorded (device will display on next wake)'})

# ==================== API: 自研6色算法处理 ====================

@app.route('/api/epd/process-sixcolor', methods=['POST'])
@login_required
def process_sixcolor():
    """使用6色算法处理图片（7.3寸E6屏）"""
    try:
        data = request.get_json(silent=True) or {}
        image_data = data.get('imageData')
        width = data.get('width', 800)
        height = data.get('height', 480)
        algorithm = data.get('algorithm', 'floyd_steinberg')
        grad_thresh = data.get('gradThresh', 40)

        if not image_data:
            return jsonify({'success': False, 'error': 'Missing imageData'}), 400

        if type(width) is not int or type(height) is not int or width != 800 or height != 480:
            return jsonify({'success': False, 'error': 'Target size must be exactly 800x480'}), 400

        if algorithm not in ['floyd_steinberg', 'gradient_blend', 'grayscale_color_map']:
            return jsonify({'success': False, 'error': f'Invalid algorithm: {algorithm}'}), 400
        if isinstance(grad_thresh, bool) or not isinstance(grad_thresh, (int, float)) or not 0 <= grad_thresh <= 255:
            return jsonify({'success': False, 'error': 'gradThresh must be between 0 and 255'}), 400

        result = process_e6_image_from_base64(
            image_data,
            width,
            height,
            algorithm=algorithm,
            grad_thresh=grad_thresh
        )
        return jsonify(result)

    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid image input'}), 400
    except Exception as e:
        print(f'❌ 6色处理错误: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Image processing failed'}), 500

# ==================== API: 模板数据代理 ====================

@app.route('/api/weather', methods=['GET'])
@login_required
def api_weather():
    """天气数据代理（避免前端跨域）"""
    city = request.args.get('city', '').strip()
    if not city:
        return jsonify({'success': False, 'error': 'Missing city parameter'}), 400
    data = _fetch_weather(city)
    if data:
        return jsonify({'success': True, 'data': data})
    return jsonify({'success': False, 'error': 'Failed to fetch weather data'}), 502


@app.route('/api/quote', methods=['GET'])
@login_required
def api_quote():
    """每日一言代理（避免前端跨域）"""
    data = _fetch_quote()
    if data:
        return jsonify({'success': True, 'data': data})
    return jsonify({'success': False, 'error': 'Failed to fetch quote'}), 502


@app.route('/api/device/template', methods=['POST'])
@login_required
def device_set_template():
    """
    为设备设置模板配置，并立即渲染保存 EPD 数据。
    如果前端提供了 imageBase64（Canvas 截图），用它做抖动处理（确保画布=预览=设备三端一致）；
    否则用后端 Pillow 渲染（定时唤醒自动更新场景）。
    同时保存 templateConfig，供设备后续唤醒时按需更新。
    """
    data = request.get_json(silent=True) or {}
    device_id = data.get('deviceId')
    template_id = data.get('templateId')
    template_config = data.get('templateConfig', {})
    if not isinstance(template_config, dict):
        return jsonify({'success': False, 'error': 'templateConfig must be an object'}), 400

    if not device_id or not template_id:
        return jsonify({'success': False, 'error': 'Missing deviceId or templateId'}), 400

    user = getattr(request, 'user', None)
    if not ensure_device_owner(device_id, user):
        return jsonify({'success': False, 'error': 'Device not found or no permission'}), 403

    clean_id = normalize_device_id(device_id)
    template_id = str(template_id).strip().lower()
    if template_id not in ACTIVE_TEMPLATE_IDS:
        return jsonify({'success': False, 'error': '当前只支持会议名牌模板'}), 400
    raw_name = str(template_config.get('name') or template_config.get('personName') or '').strip()
    normalized_config = normalize_nameplate_template_config(template_config)
    normalized_config['name'] = raw_name[:NAMEPLATE_MAX_NAME_LEN]
    template_config = normalized_config

    # ===== 渲染逻辑：前端截图优先，无截图时后端渲染 =====
    image_base64 = data.get('imageBase64')
    preview_image_b64 = None
    data_4bit_b64 = None

    if image_base64:
        # 前端提供了 Canvas 截图：用它做 Floyd-Steinberg 抖动处理
        # 确保画布（mainCanvas）= 预览（processedCanvas）= 设备屏幕 三端一致
        try:
            result = process_e6_image_from_base64(image_base64, 800, 480, algorithm='floyd_steinberg')
            epd_data = _encode_epd_string(np.array(result['colorIndices']).reshape(480, 800))
            preview_image_b64 = result.get('previewImage')
            data_4bit_b64 = result.get('data4bit')
            print(f'✅ 模板通过前端Canvas截图处理: {clean_id}, EPD长度={len(epd_data)}')
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid image input'}), 400
        except Exception as e:
            print(f'❌ 前端Canvas截图处理失败: {e}')
            return jsonify({'success': False, 'error': 'Image processing failed'}), 500
    else:
        # 无前端截图时，后端 Pillow 渲染
        try:
            render_result = render_template_with_preview(template_id, template_config)
            epd_data = render_result.get('epdData')
            if not epd_data or len(epd_data) != EPD_EXPECTED_CHARS:
                return jsonify({'success': False, 'error': 'Template rendering failed'}), 500
            preview_image_b64 = render_result.get('previewImage')
            data_4bit_b64 = render_result.get('data4bit')
            print(f'✅ 模板后端渲染完成: {clean_id}, EPD长度={len(epd_data)}')
        except Exception as e:
            print(f'❌ 模板渲染失败: {e}')
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': 'Template rendering failed'}), 500

    # 更新设备记录
    if devices_collection is not None:
        # 从 templateConfig 中读取用户自定义唤醒间隔
        custom_interval = template_config.get('sleepIntervalSeconds') if isinstance(template_config, dict) else None
        content_meta = build_content_metadata('template', template_id, custom_interval)
        now = utcnow()

        update_fields = {
            'imageSizeChars': len(epd_data),
            'imageSizeBytes': len(epd_data.encode('utf-8')),
            'imageSha256': hashlib.sha256(epd_data.encode('utf-8')).hexdigest(),
            'activeContentMode': content_meta['activeContentMode'],
            'activeContentLabel': content_meta['activeContentLabel'],
            'activeTemplateId': template_id,
            'sleepIntervalSeconds': content_meta['sleepIntervalSeconds'],
            'templateConfig': template_config,
            'renderSource': 'canvas' if image_base64 else 'pillow',
            'activeContentUpdatedAt': now,
            'updatedAt': now,
        }

        owner = user.get('username') if user else None
        with get_device_write_lock(clean_id):
            current_device = devices_collection.find_one({
                'deviceId': clean_id, 'owner': owner, 'claimed': True,
            })
            if current_device is None:
                return jsonify({'success': False, 'error': 'Device changed during publish'}), 409
            if not save_device_image(clean_id, epd_data):
                return jsonify({'success': False, 'error': 'Failed to save rendered image'}), 500
            updated_device = devices_collection.find_one_and_update(
                {'deviceId': clean_id, 'owner': owner, 'claimed': True},
                {'$inc': {'imageVersion': 1}, '$set': update_fields},
                return_document=ReturnDocument.AFTER,
            )
        if updated_device is None:
            return jsonify({'success': False, 'error': 'Device changed during publish'}), 409
        new_version = updated_device.get('imageVersion', 0)

        print(f'✅ 模板已设置并渲染: {clean_id}, 模板={template_id}, 版本={new_version}')
        response_data = {
            'success': True,
            'message': 'Template set and rendered',
            'imageVersion': new_version,
            'imageUrl': build_raw_image_url(clean_id, new_version),
            'activeContentLabel': content_meta['activeContentLabel'],
            'sleepIntervalSeconds': content_meta['sleepIntervalSeconds'],
        }
        # 返回抖动后的预览图和4bit数据，让前端更新 processedCanvas
        if preview_image_b64:
            response_data['previewImage'] = preview_image_b64
        if data_4bit_b64:
            response_data['data4bit'] = data_4bit_b64
        return jsonify(response_data)

    return jsonify({'success': True, 'message': 'Template rendered and saved'})


@app.route('/api/nameplates/dispatch', methods=['POST'])
@login_required
def dispatch_nameplates():
    """批量把人名渲染为铭牌并下发到一组设备。

    第一版只处理姓名下发；微信/AI 入口后续可以把解析后的 names 数组提交到这里。
    """
    try:
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({'success': False, 'error': 'Invalid JSON payload'}), 400
        user = getattr(request, 'user', None)
        owner = user.get('username') if user else None

        if devices_collection is None or not owner:
            return jsonify({'success': False, 'error': 'Database not connected'}), 500

        raw_names = data.get('names')
        raw_text = data.get('text') or data.get('message') or ''
        requested_device_ids = data.get('deviceIds')
        if isinstance(raw_names, list) and len(raw_names) > NAMEPLATE_MAX_NAMES:
            return jsonify({'success': False, 'error': f'姓名数量不能超过 {NAMEPLATE_MAX_NAMES} 个'}), 400
        if isinstance(raw_text, str) and len(raw_text) > NAMEPLATE_MAX_PARSE_TEXT_CHARS:
            return jsonify({'success': False, 'error': '名单文本过长'}), 400
        if (
            isinstance(requested_device_ids, list)
            and len(requested_device_ids) > NAMEPLATE_MAX_TARGET_DEVICES
        ):
            return jsonify({
                'success': False,
                'error': f'目标设备数量不能超过 {NAMEPLATE_MAX_TARGET_DEVICES} 台',
            }), 400

        names = parse_nameplate_names(data)
        if len(names) > NAMEPLATE_MAX_NAMES:
            return jsonify({'success': False, 'error': f'姓名数量不能超过 {NAMEPLATE_MAX_NAMES} 个'}), 400
        if not names:
            return jsonify({'success': False, 'error': '未识别到可下发的人名，请使用一行一个姓名或用逗号/顿号分隔'}), 400

        target_devices, missing_device_ids = resolve_nameplate_target_devices(owner, requested_device_ids)
        if missing_device_ids:
            return jsonify({
                'success': False,
                'error': '部分设备不存在或无权限',
                'missingDeviceIds': missing_device_ids
            }), 403

        if not target_devices:
            return jsonify({'success': False, 'error': '当前账号没有可下发的设备'}), 400

        base_template_config = normalize_nameplate_template_config(data.get('templateConfig'))
        custom_interval = base_template_config.get('sleepIntervalSeconds')
        content_meta = build_content_metadata('template', 'nameplate', custom_interval)
        batch_id = f"nameplate-{int(time.time())}-{secrets.token_hex(3)}"

        assignments = []
        failed = []
        assign_count = min(len(names), len(target_devices))
        dispatch_deadline = time.monotonic() + NAMEPLATE_DISPATCH_DEADLINE_SECONDS
        processed_count = 0
        deadline_reached = False

        for index in range(assign_count):
            if time.monotonic() >= dispatch_deadline:
                deadline_reached = True
                break
            processed_count = index + 1
            device = target_devices[index]
            clean_id = device.get('deviceId')
            name = names[index]
            template_config = dict(base_template_config)
            template_config['name'] = name

            try:
                render_result = render_template_with_preview('nameplate', template_config)
                epd_data = render_result.get('epdData') if isinstance(render_result, dict) else None
                if not epd_data or len(epd_data) != EPD_EXPECTED_CHARS:
                    raise ValueError(f'Nameplate rendering failed or invalid length: {len(epd_data) if epd_data else 0}')

                now = utcnow()
                active_label = f'铭牌：{name}'

                with get_device_write_lock(clean_id):
                    current_device = devices_collection.find_one({
                        'deviceId': clean_id, 'owner': owner, 'claimed': True,
                    })
                    if current_device is None:
                        raise RuntimeError('Device changed during publish')
                    if not save_device_image(clean_id, epd_data):
                        raise RuntimeError('Failed to save rendered image')
                    updated_device = devices_collection.find_one_and_update(
                        {'deviceId': clean_id, 'owner': owner, 'claimed': True},
                        {
                            '$inc': {'imageVersion': 1},
                            '$set': {
                                'imageSizeChars': len(epd_data),
                                'imageSizeBytes': len(epd_data.encode('utf-8')),
                                'imageSha256': hashlib.sha256(epd_data.encode('utf-8')).hexdigest(),
                                'activeContentMode': content_meta['activeContentMode'],
                                'activeContentLabel': active_label,
                                'activeTemplateId': 'nameplate',
                                'sleepIntervalSeconds': content_meta['sleepIntervalSeconds'],
                                'templateConfig': template_config,
                                'renderSource': 'nameplate_batch',
                                'nameplateName': name,
                                'nameplateBatchId': batch_id,
                                'activeContentUpdatedAt': now,
                                'updatedAt': now,
                            },
                        },
                        return_document=ReturnDocument.AFTER,
                    )
                if updated_device is None:
                    raise RuntimeError('Device changed during publish')
                new_version = updated_device.get('imageVersion', 0)

                assignments.append({
                    'deviceId': clean_id,
                    'deviceName': device.get('deviceName', clean_id),
                    'name': name,
                    'imageVersion': new_version,
                    'activeContentLabel': active_label,
                })
                print(f'✅ 铭牌已下发: {clean_id} -> {name}, 版本={new_version}, batch={batch_id}')
            except Exception as e:
                failed.append({
                    'deviceId': clean_id,
                    'deviceName': device.get('deviceName', clean_id),
                    'name': name,
                    'error': '处理失败',
                })
                print(f'❌ 铭牌下发失败: {clean_id} -> {name}: {e}')

        unprocessed_devices = [
            {
                'deviceId': target_devices[index].get('deviceId'),
                'deviceName': target_devices[index].get(
                    'deviceName', target_devices[index].get('deviceId')
                ),
                'name': names[index],
                'reason': '批量处理时间预算已用尽，未处理',
            }
            for index in range(processed_count, assign_count)
        ]
        if unprocessed_devices:
            deadline_reached = True
            failed.extend([
                {
                    'deviceId': item['deviceId'],
                    'deviceName': item['deviceName'],
                    'name': item['name'],
                    'error': item['reason'],
                    'unprocessed': True,
                }
                for item in unprocessed_devices
            ])

        skipped_names = names[assign_count:]
        unassigned_devices = [
            {
                'deviceId': device.get('deviceId'),
                'deviceName': device.get('deviceName', device.get('deviceId')),
            }
            for device in target_devices[assign_count:]
        ]

        if not assignments:
            return jsonify({
                'success': False,
                'error': '批量处理时间预算已用尽' if deadline_reached else '铭牌渲染或保存失败',
                'failed': failed,
                'batchId': batch_id,
                'deadlineReached': deadline_reached,
                'processedCount': processed_count,
                'unprocessedDevices': unprocessed_devices,
            }), 503 if deadline_reached else 500

        return jsonify({
            'success': True,
            'message': 'Nameplates partially dispatched' if failed else 'Nameplates dispatched',
            'batchId': batch_id,
            'assignedCount': len(assignments),
            'nameCount': len(names),
            'deviceCount': len(target_devices),
            'processedCount': processed_count,
            'deadlineReached': deadline_reached,
            'assignments': assignments,
            'failed': failed,
            'unprocessedDevices': unprocessed_devices,
            'skippedNames': skipped_names,
            'unassignedDevices': unassigned_devices,
            'sleepIntervalSeconds': content_meta['sleepIntervalSeconds'],
        })
    except Exception as e:
        print(f'❌ 批量铭牌下发异常: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': '批量下发失败'}), 500


@app.route('/api/nameplates/parse', methods=['POST'])
@login_required
def parse_nameplates():
    """把用户上传的文字/图片/表格解析成待确认的铭牌草稿。"""
    try:
        request_json = request.get_json(silent=True) or {} if request.is_json else {}
        raw_template_config = request_json.get('templateConfig', {})
        form_template_config = request.form.get('templateConfig') if request.form else None
        if form_template_config:
            try:
                raw_template_config = json.loads(form_template_config)
            except (TypeError, ValueError):
                raw_template_config = {}

        base_template_config = normalize_nameplate_template_config(
            raw_template_config
        )
        source_text, image_parts, warnings, filenames = collect_nameplate_parse_sources(request)

        if not source_text.strip() and not image_parts:
            return jsonify({'success': False, 'error': '请先输入文字或上传图片/表格'}), 400

        local_names = parse_nameplate_names_from_text(source_text)
        should_use_ai = bool(image_parts) or len(source_text) > 0
        result = None

        if should_use_ai and get_nameplate_ai_api_key():
            try:
                result = call_openai_nameplate_parser(source_text, image_parts, base_template_config)
                result['warnings'] = warnings + result.get('warnings', [])
            except Exception as e:
                warnings.append('AI解析失败，已使用本地规则解析')
                print(f'⚠️ 名单 AI 解析失败: {e}')
        elif image_parts:
            warnings.append('服务器未配置 NAMEPLATE_AI_API_KEY 或 OPENAI_API_KEY，图片内容暂不能识别')

        if result is None:
            result = build_nameplate_parse_result(
                local_names or parse_nameplate_names({'text': source_text}),
                base_template_config,
                warnings=warnings,
                ai_used=False,
                source_summary='本地规则解析',
            )

        if filenames:
            result['sourceFiles'] = filenames

        return jsonify({'success': True, 'parsed': result})
    except ValueError:
        return jsonify({'success': False, 'error': '上传内容超过限制或格式无效'}), 400
    except Exception as e:
        print(f'❌ 名单解析异常: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': '名单解析失败'}), 500


# ==================== 健康检查 ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    mongo_ok = False
    if mongo_client is not None:
        try:
            mongo_client.admin.command('ping')
            mongo_ok = True
        except Exception as e:
            print(f'❌ MongoDB health ping failed: {e}')

    response = jsonify({
        'success': mongo_ok,
        'status': 'healthy' if mongo_ok else 'degraded',
        'mongodb': 'connected' if mongo_ok else 'disconnected',
        'architecture': 'deep-sleep-http-pull',
        'mqtt': 'removed'  # 明确标注MQTT已移除
    })
    return response, 200 if mongo_ok else 503

# ==================== 启动服务器 ====================

def init_app():
    """初始化应用"""
    print('\n🚀 Starting ESP32 E-Paper Cloud Server...')
    print('📡 Architecture: Deep-sleep + HTTP Pull (No MQTT)')
    print(f'💾 MongoDB: {redact_uri_secret(Config.MONGODB_URI)} / db={Config.MONGODB_DB}')
    print(f'📁 Image Storage: {DATA_DIR}\n')

    connect_mongodb()

# 单元测试使用 mock 集合，不得连接真实 MongoDB。
if os.environ.get('APP_SKIP_INIT_FOR_TESTS', '').strip().lower() not in ('1', 'true', 'yes'):
    init_app()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f'\n🌐 API Server running on http://0.0.0.0:{port}\n')
    app.run(host='0.0.0.0', port=port, debug=Config.DEBUG)
